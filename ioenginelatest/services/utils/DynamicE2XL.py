from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from datetime import datetime as dt
import json

location = ''
try:
    location = json.load(open('/etc/autointelli/autointelli.conf', 'r'))['onapploc']
except Exception as e:
    print("Couldn't start the CRON because the location is missing in conf file")

dIPFQDNDownload = {'blr-': {'fqdn': 'r2d2.nxtgen.com'},
           'amd-': {'fqdn': 'r2d22.nxtgen.com'},
           'fbd-': {'fqdn': 'r2d23.nxtgen.com'},
           'mum-': {'fqdn': 'r2d21.nxtgen.com'}}

def export2XLSX(_2DList, pSheetName):
    try:
        try:
            # sQuery = "select am.machine_fqdn, am.ip_address, am.platform, am.osname, am.osversion, am.remediate, am.machine_id from ai_machine am left join ai_device_credentials ac on(am.fk_cred_id=ac.cred_id)"
            # sQuery = "select inventory from ai_machine where active_yn='Y'"
            # dRet = conn.returnSelectQueryResultAs2DList(sQuery)
            # dRet = conn.returnSelectQueryResult(sQuery)
            if 1 == 1:
                # xlsTmp, xlsList = dRet["data"], []
                # xlsList.append(["HOSTNAME", "IPADDRESS", "PLATFORM", "OSNAME", "VERSION", "ARCHITECTURE", "MEMORY_TOTAL", "SWAP_TOTAL", "PROCESSOR_COUNT", "DISK"])
                # for eachInv in xlsTmp:
                #    inv1 = eachInv["inventory"]
                #    xlRow = [inv1[i] for i in ["HOSTNAME", "IPADDRESS", "PLATFORM", "OSNAME", "VERSION", "ARCHITECTURE", "MEMORY_TOTAL", "SWAP_TOTAL", "PROCESSOR_COUNT"]] + ['\n'.join(["{0}: {1}".format(i, inv1["DISK"][i]) for i in list(inv1["DISK"].keys())])]
                #    xlsList.append(xlRow)

                row, sSys = 1, "linux"
                # xlsList[0] = [i.capitalize() for i in xlsList[0]]
                xlsList = _2DList
                wb = Workbook()
                ws = wb.create_sheet(pSheetName)
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                # Merged Header
                # ws.merge_cells('A1:H1')
                # ws.cell(row=1, column=1).value = "Performance Report"
                # ws['A1'].fill = PatternFill(start_color="0814FF", end_color="FFC7CE", fill_type="solid")
                # ws['A1'].font = Font(color="FFFFFF")
                # ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

                for i in xlsList:
                    ws.append(i)
                    col = 1
                    for j in i:
                        if j != '':
                            ws.cell(row=row, column=col).border = thin_border
                        col += 1
                    row += 1

                # Header
                # for eC in "ABCDEFGHIJ":
                #     ws[eC + '2'].fill = PatternFill(start_color="FFC414", end_color="FFC7CE", fill_type="solid")
                #     ws[eC + '2'].alignment = Alignment(horizontal="center", vertical="center")

                xlsxName = "{0}_".format(pSheetName) + str(int(dt.now().timestamp() * 1000000)) + ".xlsx"
                if sSys == "win":
                    xlsxPath = "E:\\" + xlsxName
                else:
                    xlsxPath = "/usr/share/nginx/html/downloads/" + xlsxName

                wb.remove(wb['Sheet'])
                wb.save(xlsxPath)
                wb.close()
                return {"result": "success", "data": "https://{0}/downloads/".format(dIPFQDNDownload[location]['fqdn']) + xlsxName}
            else:
                return {"result": "failure", "data": "Unable to download machine data"}
        except Exception as e:
            return {"result": "failure", "data": "Exception: {0}".format(str(e))}
    except Exception as e:
        return {"result": "failure", "data": "Exception: {0}".format(str(e))}