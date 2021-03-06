from services.utils.ConnLog import create_log_file
import services.utils.LFColors as lfc
import json
from services.utils.ValidatorSession import chkValidRequest, chkKeyExistsInHeader, lam_api_key_invalid, lam_api_key_missing
from flask import request
from elasticsearch import Elasticsearch
from datetime import datetime as dt
import pytz
import math
from copy import deepcopy
import services.utils.ConnPostgreSQL as pcon
import requests as req
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from services.perfrpt.ArrayLBRS import getRealServices

lfcObj = lfc.bcolors()
CERROR, CWARN = lfcObj.printerr, lfcObj.printwar
logObj = create_log_file()
if not logObj:
    CERROR("Not able to create logfile")
    exit(0)
logERROR, logWARN, logINFO = logObj.error, logObj.warn, logObj.info

location = ''
try:
    location = json.load(open('/etc/autointelli/autointelli.conf', 'r'))['onapploc']
except Exception as e:
    print("Couldn't start the CRON because the location is missing in conf file")

dIPFQDN = {'blr-': [{'host': '10.227.45.119', 'port': 9200}, {'host': '10.227.45.120', 'port': 9200}, {'host': '10.227.45.121', 'port': 9200}],
           'amd-': [{'host': '10.210.45.119', 'port': 9200}, {'host': '10.210.45.120', 'port': 9200}, {'host': '10.210.45.121', 'port': 9200}],
           'fbd-': [{'host': '10.195.45.119', 'port': 9200}, {'host': '10.195.45.120', 'port': 9200}, {'host': '10.195.45.121', 'port': 9200}],
           'mum-': [{'host': '10.239.45.218', 'port': 9200}, {'host': '10.239.45.219', 'port': 9200}, {'host': '10.239.45.220', 'port': 9200}, {'host': '10.239.45.221', 'port': 9200}, {'host': '10.239.45.222', 'port': 9200}, {'host': '10.239.45.223', 'port': 9200}]}

dIPFQDNDownload = {'blr-': {'fqdn': 'r2d2.nxtgen.com'},
           'amd-': {'fqdn': '61.0.172.106'},
           'fbd-': {'fqdn': '117.255.216.170'},
           'mum-': {'fqdn': '103.230.37.88'}}

def logAndRet(s, x):
    if s == "failure":
        logERROR(x)
    else:
        logINFO(x)
    return json.dumps({"result": s, "data": x})

def perfData(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                _index = "ai-network-bandwidth"
                timezone = dPayload["TimeZone"] #"Asia/Kolkata"
                start_datetime = dPayload["start_datetime"] #"2020-08-29 00:00"
                end_datetime = dPayload["end_datetime"] #"2020-08-30 00:00"
                ip = dPayload["ip"] #"10.225.11.102"
                port = dPayload["port"] #"54001"
                user_tz = pytz.timezone(timezone)
                es_tz = pytz.timezone('GMT')

                user_s, user_e = user_tz.localize(dt.strptime(start_datetime, '%Y-%m-%d %H:%M')), user_tz.localize(
                    dt.strptime(end_datetime, '%Y-%m-%d %H:%M'))  # Declare I/P to particular timezone
                gmt_s, gmt_e = user_s.astimezone(es_tz), user_e.astimezone(es_tz)
                start_datetime = gmt_s.strftime('%Y-%m-%dT%H:%MZ')
                end_datetime = gmt_e.strftime('%Y-%m-%dT%H:%MZ')
                _bucket = math.ceil(((gmt_e.timestamp() - gmt_s.timestamp()) / 60) / 1000)

                body = {
                    "query": {
                        "bool": {
                            "filter": [
                                {
                                    "match": {
                                        "Hostname": ip

                                        #"{0}.{1}.Name".format(ip, port): "Ethernet54/1"
                                    }
                                },
                                {
                                    "match":{
                                        "Interface": port
                                    }

                                },
                                {
                                    "range": {
                                        "@timestamp.GMT": {
                                            "gte": start_datetime,
                                            "lte": end_datetime
                                        }
                                    }
                                }
                            ]
                        }
                    }
                }
                body['aggs'] = {'Date': {'date_histogram': {'field': '@timestamp.GMT', 'fixed_interval': '{0}m'.format(_bucket)},
                                           'aggs': {}}}
                #body['size'] = 0
                d = {}
                for i in ['Traffic Total(speed)', 'Traffic In(speed)', 'Traffic Out(speed)', 'Traffic Total(volume)', 'Traffic In(volume)', 'Traffic Out(volume)']:
                    d[i] = {'avg': {'field': i}}
                body['aggs']['Date']['aggs'] = d

                print(body)
                es = Elasticsearch(dIPFQDN[location])
                result = es.search(index=_index, scroll='2m', body=body, size=10000)

                FinalData, ft = {}, True
                name, desc, speed, volume = "", "", "", ""
                sid = result['_scroll_id']
                scroll_size = len(result['hits']['hits'])
                tmp = {}
                if scroll_size <= 0:
                    return json.dumps({"result": "failure", "data": "no data"})
                while scroll_size > 0:
                    data = es.scroll(scroll_id=sid, scroll='2m')
                    for d in result['aggregations']['Date']['buckets']:
                        print(d)
                        #d = i['_source']#[ip][port]
                        dtime = es_tz.localize(
                            dt.strptime(d['key_as_string'], '%Y-%m-%dT%H:%M:%S.%fZ')).astimezone(user_tz)
                        tmp[dtime] = [float('%.2f' % d['Traffic In(volume)']['value']), float('%.2f' % d['Traffic Out(volume)']['value']), float('%.2f' % d['Traffic Total(volume)']['value']),float('%.2f' % d['Traffic In(speed)']['value']), float('%.2f' % d['Traffic Out(speed)']['value']), float('%.2f' % d['Traffic Total(speed)']['value'])]
                        # tmp.append([dtime.strftime('%Y-%m-%d %H:%M'), d['Traffic In(volume)'], d['Traffic Out(volume)'], d['Traffic Total(volume)'],
                        #          d['Traffic In(speed)'], d['Traffic Out(speed)'], d['Traffic Total(speed)']])
                        if ft:
                            name = result['hits']['hits'][0]['_source']['Name']
                            desc = result['hits']['hits'][0]['_source']['Description']
                            speed = result['hits']['hits'][0]['_source']['speed(metric)']
                            volume = result['hits']['hits'][0]['_source']['volume(metric)']
                            ft = False
                    sid = data['_scroll_id']
                    scroll_size = len(data['hits']['hits'])

                FinalData['Name'] = name
                FinalData['Description'] = desc
                FinalData['SpeedMetrics'] = speed
                FinalData['VolumeMetrics'] = volume
                finalList = []
                for i in sorted(tmp):
                    x = [i.strftime('%Y-%m-%d %H:%M')] + tmp[i]
                    finalList.append(x)
                finalList.insert(0, ['Date Time', 'Traffic In(volume)', 'Traffic Out(volume)', 'Traffic Total(volume)',
                                     'Traffic In(speed)', 'Traffic Out(speed)', 'Traffic Total(speed)'])
                FinalData['plots'] = finalList

                # Aggregate
                # aggFinal = []
                # aggre = 5
                # agg = deepcopy(FinalData['plots'][1:])
                # inv, outv, totv, ins, outs, tots, mid = 0, 0, 0, 0, 0, 0, ""
                # mid = agg
                # for i in range(0, len(mid) - 1, aggre):
                #     for j in range(0, aggre):
                #         inv += float(mid[i + j][1])
                #         outv += float(mid[i + j][2])
                #         totv += float(mid[i + j][3])
                #         ins += float(mid[i + j][4])
                #         outs += float(mid[i + j][5])
                #         tots += float(mid[i + j][6])
                #     aggFinal.append(
                #         ["{0} - {1}".format(mid[i][0], mid[i + (aggre-1)][0]), math.ceil(inv), math.ceil(outv),
                #          math.ceil(totv), math.ceil(ins / aggre), math.ceil(outs / aggre),
                #          math.ceil(tots / aggre)])
                # aggFinal.insert(0, ['Date Time', 'Traffic In(volume)', 'Traffic Out(volume)', 'Traffic Total(volume)',
                #                     'Traffic In(speed)', 'Traffic Out(speed)', 'Traffic Total(speed)'])
                # FinalData['plots'] = aggFinal

                # Aggregate
                # aggFinal = []
                # aggre = 5
                # agg = deepcopy(FinalData['plots'][1:])
                # iniInd = int(dt.strptime(agg[0][0], '%Y-%m-%d %H:%M').strftime('%M')) % aggre
                # finInd = int(dt.strptime(agg[-1][0], '%Y-%m-%d %H:%M').strftime('%M')) % aggre
                #
                # if iniInd > 0:
                #     inv, outv, totv, ins, outs, tots = 0, 0, 0, 0, 0, 0
                #     first = agg[0:iniInd]
                #     samples = len(first)
                #     for i in first:
                #         inv += float(i[1])
                #         outv += float(i[2])
                #         totv += float(i[3])
                #         ins += float(i[4])
                #         outs += float(i[5])
                #         tots += float(i[6])
                #     aggFinal.append(["{0} - {1}".format(first[0][0], first[-1][0]), math.ceil(inv), math.ceil(outv),
                #                      math.ceil(totv), math.ceil(ins / samples), math.ceil(outs / samples),
                #                      math.ceil(tots / samples)])
                #
                # inv, outv, totv, ins, outs, tots, mid = 0, 0, 0, 0, 0, 0, ""
                # if (iniInd == 0 and finInd == 0) or finInd == 0:
                #     mid = agg[iniInd:]
                # else:
                #     mid = agg[iniInd:-finInd]
                # for i in range(0, len(mid)-1, aggre):
                #     for j in range(0, aggre):
                #         inv += float(mid[i + j][1])
                #         outv += float(mid[i + j][2])
                #         totv += float(mid[i + j][3])
                #         ins += float(mid[i + j][4])
                #         outs += float(mid[i + j][5])
                #         tots += float(mid[i + j][6])
                #     aggFinal.append(
                #         ["{0} - {1}".format(mid[i][0], mid[i + aggre][0]), math.ceil(inv), math.ceil(outv),
                #          math.ceil(totv), math.ceil(ins / aggre), math.ceil(outs / aggre),
                #          math.ceil(tots / aggre)])
                #
                # if finInd > 0:
                #     inv, outv, totv, ins, outs, tots = 0, 0, 0, 0, 0, 0
                #     last = agg[-finInd:]
                #     samples = len(last)
                #     for i in last:
                #         inv += float(i[1])
                #         outv += float(i[2])
                #         totv += float(i[3])
                #         ins += float(i[4])
                #         outs += float(i[5])
                #         tots += float(i[6])
                #     aggFinal.append(["{0} - {1}".format(last[0][0], last[-1][0]), math.ceil(inv), math.ceil(outv), math.ceil(totv),
                #          math.ceil(ins / samples), math.ceil(outs / samples), math.ceil(tots / samples)])
                #
                # aggFinal.insert(0, ['Date Time', 'Traffic In(volume)', 'Traffic Out(volume)', 'Traffic Total(volume)',
                #                      'Traffic In(speed)', 'Traffic Out(speed)', 'Traffic Total(speed)'])
                # FinalData['plots'] = aggFinal

                # Sum and Average
                samples = len(FinalData['plots'][1:])
                vin, vout, vtot, sin, sout, stot = 0, 0, 0, 0, 0, 0
                for i in FinalData['plots'][1:]:
                    vin += float(i[1])
                    vout += float(i[2])
                    vtot += float(i[3])
                    sin += float(i[4])
                    sout += float(i[5])
                    stot += float(i[6])
                FinalData['Sums'] = {'Traffic In(volume)': float('%.2f' % math.ceil(vin)), 'Traffic Out(volume)': float('%.2f' % math.ceil(vout)), 'Traffic Total(volume)': float('%.2f' % math.ceil(vtot))}
                FinalData['Average'] = {'Traffic In(volume)': float('%.2f' % math.ceil(vin/samples)), 'Traffic Out(volume)': float('%.2f' % math.ceil(vout/samples)), 'Traffic Total(volume)': float('%.2f' % math.ceil(vtot/samples)),
                                      'Traffic In(speed)': float('%.2f' % math.ceil(sin/samples)), 'Traffic Out(speed)': float('%.2f' % math.ceil(sout/samples)), 'Traffic Total(speed)': float('%.2f' % math.ceil(stot/samples))}
                return json.dumps({"result": "success", "data": FinalData})
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def perfData1(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                _index = "ai-network-bandwidth"
                timezone = dPayload["TimeZone"] #"Asia/Kolkata"
                start_datetime = dPayload["start_datetime"] #"2020-08-29 00:00"
                end_datetime = dPayload["end_datetime"] #"2020-08-30 00:00"
                ip = dPayload["ip"] #"10.225.11.102"
                port = dPayload["port"] #"54001"
                user_tz = pytz.timezone(timezone)
                es_tz = pytz.timezone('GMT')

                user_s, user_e = user_tz.localize(dt.strptime(start_datetime, '%Y-%m-%d %H:%M')), user_tz.localize(
                    dt.strptime(end_datetime, '%Y-%m-%d %H:%M'))  # Declare I/P to particular timezone
                gmt_s, gmt_e = user_s.astimezone(es_tz), user_e.astimezone(es_tz)
                start_datetime = gmt_s.strftime('%Y-%m-%dT%H:%MZ')
                end_datetime = gmt_e.strftime('%Y-%m-%dT%H:%MZ')

                body = {
                    "query": {
                        "bool": {
                            "filter": [
                                {
                                    "match": {
                                        "Hostname": ip

                                        #"{0}.{1}.Name".format(ip, port): "Ethernet54/1"
                                    }
                                },
                                {
                                    "match":{
                                        "Interface": port
                                    }

                                },
                                {
                                    "range": {
                                        "@timestamp.GMT": {
                                            "gte": start_datetime,
                                            "lte": end_datetime
                                        }
                                    }
                                }
                            ]
                        }
                    }
                }

                print(body)
                es = Elasticsearch(dIPFQDN[location])
                result = es.search(index=_index, scroll='2m', body=body, size=10000)

                FinalData, ft = {}, True
                name, desc, speed, volume = "", "", "", ""
                sid = result['_scroll_id']
                scroll_size = len(result['hits']['hits'])
                tmp = {}
                if scroll_size <= 0:
                    return json.dumps({"result": "failure", "data": "no data"})
                while scroll_size > 0:
                    data = es.scroll(scroll_id=sid, scroll='2m')
                    for i in result['hits']['hits']:
                        d = i['_source']#[ip][port]
                        dtime = es_tz.localize(
                            dt.strptime(i['_source']['@timestamp']['GMT'], '%Y-%m-%dT%H:%M:%S.%fZ')).astimezone(user_tz)
                        tmp[dtime] = [d['Traffic In(volume)'], d['Traffic Out(volume)'], d['Traffic Total(volume)'],
                                      d['Traffic In(speed)'], d['Traffic Out(speed)'], d['Traffic Total(speed)']]
                        # tmp.append([dtime.strftime('%Y-%m-%d %H:%M'), d['Traffic In(volume)'], d['Traffic Out(volume)'], d['Traffic Total(volume)'],
                        #          d['Traffic In(speed)'], d['Traffic Out(speed)'], d['Traffic Total(speed)']])
                        if ft:
                            name = d['Name']
                            desc = d['Description']
                            speed = d['speed(metric)']
                            volume = d['volume(metric)']
                            ft = False
                    sid = data['_scroll_id']
                    scroll_size = len(data['hits']['hits'])

                FinalData['Name'] = name
                FinalData['Description'] = desc
                FinalData['SpeedMetrics'] = speed
                FinalData['VolumeMetrics'] = volume
                finalList = []
                for i in sorted(tmp):
                    x = [i.strftime('%Y-%m-%d %H:%M')] + tmp[i]
                    finalList.append(x)
                finalList.insert(0, ['Date Time', 'Traffic In(volume)', 'Traffic Out(volume)', 'Traffic Total(volume)',
                                     'Traffic In(speed)', 'Traffic Out(speed)', 'Traffic Total(speed)'])
                FinalData['plots'] = finalList

                # Aggregate
                # aggFinal = []
                # aggre = 5
                # agg = deepcopy(FinalData['plots'][1:])
                # inv, outv, totv, ins, outs, tots, mid = 0, 0, 0, 0, 0, 0, ""
                # mid = agg
                # for i in range(0, len(mid) - 1, aggre):
                #     for j in range(0, aggre):
                #         inv += float(mid[i + j][1])
                #         outv += float(mid[i + j][2])
                #         totv += float(mid[i + j][3])
                #         ins += float(mid[i + j][4])
                #         outs += float(mid[i + j][5])
                #         tots += float(mid[i + j][6])
                #     aggFinal.append(
                #         ["{0} - {1}".format(mid[i][0], mid[i + (aggre-1)][0]), math.ceil(inv), math.ceil(outv),
                #          math.ceil(totv), math.ceil(ins / aggre), math.ceil(outs / aggre),
                #          math.ceil(tots / aggre)])
                # aggFinal.insert(0, ['Date Time', 'Traffic In(volume)', 'Traffic Out(volume)', 'Traffic Total(volume)',
                #                     'Traffic In(speed)', 'Traffic Out(speed)', 'Traffic Total(speed)'])
                # FinalData['plots'] = aggFinal

                # Aggregate
                # aggFinal = []
                # aggre = 5
                # agg = deepcopy(FinalData['plots'][1:])
                # iniInd = int(dt.strptime(agg[0][0], '%Y-%m-%d %H:%M').strftime('%M')) % aggre
                # finInd = int(dt.strptime(agg[-1][0], '%Y-%m-%d %H:%M').strftime('%M')) % aggre
                #
                # if iniInd > 0:
                #     inv, outv, totv, ins, outs, tots = 0, 0, 0, 0, 0, 0
                #     first = agg[0:iniInd]
                #     samples = len(first)
                #     for i in first:
                #         inv += float(i[1])
                #         outv += float(i[2])
                #         totv += float(i[3])
                #         ins += float(i[4])
                #         outs += float(i[5])
                #         tots += float(i[6])
                #     aggFinal.append(["{0} - {1}".format(first[0][0], first[-1][0]), math.ceil(inv), math.ceil(outv),
                #                      math.ceil(totv), math.ceil(ins / samples), math.ceil(outs / samples),
                #                      math.ceil(tots / samples)])
                #
                # inv, outv, totv, ins, outs, tots, mid = 0, 0, 0, 0, 0, 0, ""
                # if (iniInd == 0 and finInd == 0) or finInd == 0:
                #     mid = agg[iniInd:]
                # else:
                #     mid = agg[iniInd:-finInd]
                # for i in range(0, len(mid)-1, aggre):
                #     for j in range(0, aggre):
                #         inv += float(mid[i + j][1])
                #         outv += float(mid[i + j][2])
                #         totv += float(mid[i + j][3])
                #         ins += float(mid[i + j][4])
                #         outs += float(mid[i + j][5])
                #         tots += float(mid[i + j][6])
                #     aggFinal.append(
                #         ["{0} - {1}".format(mid[i][0], mid[i + aggre][0]), math.ceil(inv), math.ceil(outv),
                #          math.ceil(totv), math.ceil(ins / aggre), math.ceil(outs / aggre),
                #          math.ceil(tots / aggre)])
                #
                # if finInd > 0:
                #     inv, outv, totv, ins, outs, tots = 0, 0, 0, 0, 0, 0
                #     last = agg[-finInd:]
                #     samples = len(last)
                #     for i in last:
                #         inv += float(i[1])
                #         outv += float(i[2])
                #         totv += float(i[3])
                #         ins += float(i[4])
                #         outs += float(i[5])
                #         tots += float(i[6])
                #     aggFinal.append(["{0} - {1}".format(last[0][0], last[-1][0]), math.ceil(inv), math.ceil(outv), math.ceil(totv),
                #          math.ceil(ins / samples), math.ceil(outs / samples), math.ceil(tots / samples)])
                #
                # aggFinal.insert(0, ['Date Time', 'Traffic In(volume)', 'Traffic Out(volume)', 'Traffic Total(volume)',
                #                      'Traffic In(speed)', 'Traffic Out(speed)', 'Traffic Total(speed)'])
                # FinalData['plots'] = aggFinal

                # Sum and Average
                samples = len(FinalData['plots'][1:])
                vin, vout, vtot, sin, sout, stot = 0, 0, 0, 0, 0, 0
                for i in FinalData['plots'][1:]:
                    vin += float(i[1])
                    vout += float(i[2])
                    vtot += float(i[3])
                    sin += float(i[4])
                    sout += float(i[5])
                    stot += float(i[6])
                FinalData['Sums'] = {'Traffic In(volume)': math.ceil(vin), 'Traffic Out(volume)': math.ceil(vout), 'Traffic Total(volume)': math.ceil(vtot)}
                FinalData['Average'] = {'Traffic In(volume)': math.ceil(vin/samples), 'Traffic Out(volume)': math.ceil(vout/samples), 'Traffic Total(volume)': math.ceil(vtot/samples),
                                      'Traffic In(speed)': math.ceil(sin/samples), 'Traffic Out(speed)': math.ceil(sout/samples), 'Traffic Total(speed)': math.ceil(stot/samples)}
                return json.dumps({"result": "success", "data": FinalData})
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def perfMaster(user_id):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                data = ""
                if user_id == "tommyuser":
                    data = {'10.225.11.20':{
                        '380': 'TommyHill-3021'
                        }
                    }
                else:
                    data = {'10.225.11.102':
                                 {'2': 'NG4-UPLINK-FW-1500D-PORT-40',
                                  '49001': 'Connected-to-spine-1-fo-1/0/49',
                                  '50001': 'Connected-to-spine-2-fo-2/0/50',
                                  '51001': 'ARISTA-BMS-R1',
                                  '53001': 'Check Uplink Traffic',
                                  '54001': 'New-Airtel-40-G-Link Traffic'},
                             '10.225.11.103':
                                 {'1': 'ET1',
                                  '2': 'NG4-UPLINK-FW-1500D-PORT-39',
                                  '43': 'TCL-2G-ILL-Circuit-ID_091BENG030029471009',
                                  '44': 'ET44',
                                  '47': 'Everbiz-FW',
                                  '49001': 'Connected-to-spine-2-fo-2/0/49',
                                  '50001': 'Connected-to-spine-1-fo-1/0/50'},
                             '10.225.11.20':
                                {'60': 'BlackMount_3003',
                                 '30': 'BlackMount_3013',
                                 '87': 'Embassy-3163',
                                 '95': 'Embassy-3180',
                                 '73': 'jklakshmi-3154',
                                 '81': 'Tenshi-SOVI-WAN',
                                 '80': 'Tenshi-WAN',
                                 '380': 'TommyHill-3021'
                                },
                             '10.225.179.129':
                                {'666': 'Modicare-2844',
                                 '1129': 'NB_Interface-1129',
                                 '641': 'PeopleStron-WAN',
                                 '639': 'Aujas-WAN'
                                },
                            '10.225.11.12':
                                {'22': 'Uplink_LAN',
                                '23': 'Uplink_WAN',
                                '27': 'SnapD_WAN',
                                '28': 'FW-to-L3',
                                '35': 'P2P-BID-MUM'
                                }
                             }
                return json.dumps({"result": "success", "data": data})
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def perfDropDownItemMaster():
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                final = {}
                sQueryvCenter = """select A.vcenter || '..' || A.dc || '..' || B.objects objects from (
select hypervisor_ip_address vCenter, object_name DC from vcenter_object_inventory i inner join hypervisor_details h on(i.fk_hypervisor_id=h.pk_hypervisor_id) where i.object_type='datacenter') A, 
(SELECT x.objects FROM (VALUES ('ESXi Host'::text), ('ESXi VM'::text), ('Datastore'::text))  x(objects)) B order by A.vcenter, A.dc, B.objects"""
                sQueryKVM = """select * from (SELECT x.objects FROM (VALUES ('KVM Host'::text), ('KVM VM'::text))  x(objects)) B"""
                retvc = pcon.returnSelectQueryResultAsList(sQueryvCenter)
                retk = pcon.returnSelectQueryResultAsList(sQueryKVM)
                if retvc['result'] == 'success' and retk['result'] == 'success':
                    final['vCenter'] = retvc['data']['objects']
                    final['KVM'] = retk['data']['objects']
                    final['Firewall'] = ''
                    final['Switch'] = ''
                    final['Router'] = ''
                    final['Load Balancer'] = ['10.210.11.184']
                    return json.dumps({'result': 'success', 'data': final})
                else:
                    return json.dumps({'result': 'failure', 'data': 'no data'})
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getperfDropDownList4SelectedItem(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sFinal = {}
                metrics = {
                    'esxivm': ['CPU', 'Memory', 'NIC'], # 'Disk',
                    'esxihost': ['CPU', 'Memory', 'Disk', 'NIC'],
                    'datastore': ['Disk'],
                    'kvmhost': ['CPU', 'Memory', 'Disk', 'NIC'],
                    'kvmvm': ['CPU', 'Disk', 'NIC']
                }
                abbr = {'ESXi Host': 'esxihost', 'ESXi VM': 'esxivm', 'Datastore': 'datastore', 'KVM Host': 'kvmhost', 'KVM VM': 'kvmvm'}
                category = dPayload['category']
                item = dPayload['item']
                sQuery = ""

                # vCenter
                if category.lower().strip() == 'vcenter':
                    vc, dc, item1 = item.split('..')[0], item.split('..')[1], item.split('..')[2]
                    if abbr[item1] == 'esxivm':
                        sQuery = """
                        select distinct 
                        	vcenti.pk_object_id, 
                        	vcenti.object_name,
                        	vcloi.customer_id || '::' || vcloi.customer_name customer,
                        	vcents.object_state 
                        from 
                        	vcenter_object_inventory vcenti
                        	left join vcloud_object_inventory vcloi on(vcenti.object_ref=vcloi.vm_vcenter_ref) 
                        	left join vcenter_object_state vcents on(vcenti.object_ref=vcents.object_ref) 
                        where 
                        	vcenti.dc_id=(select pk_object_id from vcenter_object_inventory where fk_hypervisor_id=(select pk_hypervisor_id from hypervisor_details where hypervisor_ip_address='{0}') and object_name='{1}') and 
                        	vcenti.object_type='{2}'
                        order by 
                        	vcenti.object_name""".format(vc, dc, abbr[item1])
                        sQuery = """
select 
	distinct vcenti.pk_object_id, vcenti.object_name, vcloi.customer_id || '::' || vcloi.customer_name customer, vcents.object_state 
from 
	vcenter_object_inventory vcenti 
	left join vcloud_object_inventory vcloi on(vcenti.object_ref=vcloi.vm_vcenter_ref) 
	left join ( select * from  vcenter_object_state where dc_id=(select pk_object_id from vcenter_object_inventory where fk_hypervisor_id=(select pk_hypervisor_id from hypervisor_details where hypervisor_ip_address='{0}') and object_name='{1}') ) vcents on(vcenti.object_ref=vcents.object_ref) 
where 
	vcenti.dc_id=(select pk_object_id from vcenter_object_inventory where fk_hypervisor_id=(select pk_hypervisor_id from hypervisor_details where hypervisor_ip_address='{2}') and object_name='{3}') and 
	vcenti.object_type='{4}'
order by 
	vcenti.object_name""".format(vc, dc, vc, dc, abbr[item1])
                    else:
                        sQuery = """
select 
	pk_object_id, object_name 
from 
	vcenter_object_inventory 
where 
	dc_id=(select pk_object_id from vcenter_object_inventory where fk_hypervisor_id=(select pk_hypervisor_id from hypervisor_details where hypervisor_ip_address='{0}') and object_name='{1}') and object_type='{2}'  
order by object_name""".format(vc, dc, abbr[item1])
                    dRet = pcon.returnSelectQueryResultAs2DList(sQuery)
                    if dRet['result'] == 'success':
                        sFinal['items'] = dRet['data']
                        sFinal['metrics'] = metrics[abbr[item1]]
                        sFinal['metrics_control'] = 'radio'
                        return json.dumps({'result': 'success', 'data': sFinal})
                    else:
                        return json.dumps({'result': 'failure', 'data': 'no data'})

                # KVM
                elif category.lower().strip() == 'kvm':
                    if abbr[item] == 'kvmhost':
                        sQuery = """
select 
	distinct h_id host_id, h_ip, h_host host, h_label hlabel
from 
	onapp_object_inventory 
where active_yn='Y' 
order by h_label"""
                    elif abbr[item] == 'kvmvm':
                        sQuery = """
select 
	v_identifier VM_ID, v_label vm_name, v_operating_system VM_OS, c_login customer_id, v_ip_addresses vm_IP 
from 
	onapp_object_inventory 
where active_yn='Y' 
order by v_label"""
                    dRet = pcon.returnSelectQueryResultAs2DList(sQuery)
                    if dRet['result'] == 'success':
                        sFinal['items'] = dRet['data']
                        sFinal['metrics'] = metrics[abbr[item]]
                        sFinal['metrics_control'] = 'radio'
                        return json.dumps({'result': 'success', 'data': sFinal})
                    else:
                        return json.dumps({'result': 'failure', 'data': 'no data'})

                # Firewall
                elif category.lower().strip() == 'firewall':
                    sQuery = "select pk_hypervisor_id firewall_id, hypervisor_ip_address firewall_ip_address from hypervisor_details where lower(trim(hypervisor_type)) = 'firewall'"
                    dRet = pcon.returnSelectQueryResultAs2DList(sQuery)
                    if dRet['result'] == 'success':
                        sFinal['items'] = dRet['data']
                        sFinal['metrics'] = 'NA'
                        sFinal['metrics_control'] = 'NA'
                        return json.dumps({'result': 'success', 'data': sFinal})
                    else:
                        return json.dumps({'result': 'failure', 'data': 'no data'})

                # Switch
                elif category.lower().strip() == 'switch':
                    sQuery = "select pk_hypervisor_id switch_id, hypervisor_ip_address switch_ip_address from hypervisor_details where lower(trim(hypervisor_type)) = 'switch'"
                    dRet = pcon.returnSelectQueryResultAs2DList(sQuery)
                    if dRet['result'] == 'success':
                        sFinal['items'] = dRet['data']
                        sFinal['metrics'] = 'NA'
                        sFinal['metrics_control'] = 'NA'
                        return json.dumps({'result': 'success', 'data': sFinal})
                    else:
                        return json.dumps({'result': 'failure', 'data': 'no data'})

                # LB
                elif category.lower().strip() == 'load balancer':
                    retLB = getRealServices(item)
                    if retLB['result'] == 'success':
                        sFinal["items"] = retLB["data"]["items"]
                        sFinal["metrics"] = retLB["data"]["metrics"]
                        sFinal['metrics_control'] = 'NA'
                        return json.dumps({'result': 'success', 'data': sFinal})
                    else:
                        return json.dumps({'result': 'failure', 'data': 'no data'})

            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getperfDropDownList4SelectedItemNetwork(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                final = {}
                category = dPayload['category']
                ip = dPayload['ip']
                sQuery = ""
                if category.lower().strip() == 'firewall':
                    sQuery = """select int_id || '..' || int_name interface from firewall_object_inventory where fk_hypervisor_id=(select pk_hypervisor_id from hypervisor_details where lower(trim(hypervisor_type)) = 'firewall' and hypervisor_ip_address='{0}')""".format(
                        ip
                    )
                elif category.lower().strip() == 'switch':
                    sQuery = """select int_id || '..' || int_name interface from switch_object_inventory where fk_hypervisor_id=(select pk_hypervisor_id from hypervisor_details where lower(trim(hypervisor_type)) = 'switch' and hypervisor_ip_address='{0}')""".format(
                        ip
                    )
                dRet = pcon.returnSelectQueryResultAsList(sQuery)
                if dRet['result'] == 'success':
                    final['metrics'] = dRet['data']['interface']
                    final['metrics_control'] = 'dropdown'
                    return json.dumps({'result': 'success', 'data': final})
                else:
                    return json.dumps({'result': 'failure', 'data': 'no data'})
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def export2XLSX(_2DList):
    try:
        try:
            # sQuery = "select am.machine_fqdn, am.ip_address, am.platform, am.osname, am.osversion, am.remediate, am.machine_id from ai_machine am left join ai_device_credentials ac on(am.fk_cred_id=ac.cred_id)"
            # sQuery = "select inventory from ai_machine where active_yn='Y'"
            # dRet = conn.returnSelectQueryResultAs2DList(sQuery)
            # dRet = conn.returnSelectQueryResult(sQuery)
            if 1 == 1:
                # xlsTmp, xlsList = dRet["data"], []
                # xlsList.append(["HOSTNAME", "IPADDRESS", "PLATFORM", "OSNAME", "VERSION", "ARCHITECTURE", "MEMORY_TOTAL", "SWAP_TOTAL", "PROCESSOR_COUNT", "DISK"])
                # for eachInv in xlsTmp:
                #    inv1 = eachInv["inventory"]
                #    xlRow = [inv1[i] for i in ["HOSTNAME", "IPADDRESS", "PLATFORM", "OSNAME", "VERSION", "ARCHITECTURE", "MEMORY_TOTAL", "SWAP_TOTAL", "PROCESSOR_COUNT"]] + ['\n'.join(["{0}: {1}".format(i, inv1["DISK"][i]) for i in list(inv1["DISK"].keys())])]
                #    xlsList.append(xlRow)

                row, sSys = 2, "linux"
                # xlsList[0] = [i.capitalize() for i in xlsList[0]]
                xlsList = _2DList
                wb = Workbook()
                ws = wb.create_sheet("Performance Report")
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                # Merged Header
                ws.merge_cells('A1:H1')
                ws.cell(row=1, column=1).value = "Performance Report"
                ws['A1'].fill = PatternFill(start_color="0814FF", end_color="FFC7CE", fill_type="solid")
                ws['A1'].font = Font(color="FFFFFF")
                ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

                for i in xlsList:
                    ws.append(i)
                    col = 1
                    for j in i:
                        if j != '':
                            ws.cell(row=row, column=col).border = thin_border
                        col += 1
                    row += 1

                # Header
                # for eC in "ABCDEFGHIJ":
                #     ws[eC + '2'].fill = PatternFill(start_color="FFC414", end_color="FFC7CE", fill_type="solid")
                #     ws[eC + '2'].alignment = Alignment(horizontal="center", vertical="center")

                xlsxName = "PerformanceReport_" + str(int(dt.now().timestamp() * 1000000)) + ".xlsx"
                if sSys == "win":
                    xlsxPath = "E:\\" + xlsxName
                else:
                    xlsxPath = "/usr/share/nginx/html/downloads/" + xlsxName

                wb.remove(wb['Sheet'])
                wb.save(xlsxPath)
                wb.close()
                return {"result": "success", "data": "https://{0}/downloads/".format(dIPFQDNDownload[location]['fqdn']) + xlsxName}
            else:
                return {"result": "failure", "data": "Unable to download machine data"}
        except Exception as e:
            return {"result": "failure", "data": "Exception: {0}".format(str(e))}
    except Exception as e:
        return {"result": "failure", "data": "Exception: {0}".format(str(e))}

def esExecWithAgg(_pindex, _pbody, _es_tz, _user_tz, _bucket):
    try:
        es = Elasticsearch(dIPFQDN[location])
        result = es.search(index=_pindex, scroll='2m', body=_pbody, size=10000)
        FinalData, ft = [], True
        sid = result['_scroll_id']
        scroll_size = len(result['hits']['hits'])
        if scroll_size <= 0:
            return {"result": "failure", "data": "no data"}
        while scroll_size > 0:
            data = es.scroll(scroll_id=sid, scroll='{0}m'.format(_bucket))
            for i in result['hits']['hits']:
                FinalData.append(i['_source'])
            sid = data['_scroll_id']
            scroll_size = len(data['hits']['hits'])
        return {'result': 'success', 'data': FinalData}
    except Exception as e:
        print(str(e))
        return {'result': 'failure', 'data': []}

def makeZero(n):
    try:
        n = float(n)
    except Exception as e:
        pass
    if type(n) == type(1) or type(n) == type(1.0):
        if n < 0:
            return 0
        else:
            return n
    else:
        return n

def esExec(_pindex, _pbody, _es_tz, _user_tz, _bucket, _units, dtype = 'null'):
    try:
        es = Elasticsearch(dIPFQDN[location])
        result = es.search(index=_pindex, scroll='2m', body=_pbody, size=1000)
        FinalData, ft = [], True
        _sum, _avg = {}, {}
        _lowhigh = {}
        items = _pbody['_source'][1:]
        sid = result['_scroll_id']
        scroll_size = len(result['hits']['hits'])
        if scroll_size <= 0:
            return {"result": "failure", "data": "no data"}
        while scroll_size > 0:
            data = es.scroll(scroll_id=sid, scroll='{0}m'.format(_bucket))
            for i in result['aggregations']['Date']['buckets']:
                _date = _es_tz.localize(dt.strptime(i['key_as_string'], '%Y-%m-%dT%H:%M:%S.%fZ')).astimezone(_user_tz).strftime('%Y-%m-%d %H:%M')
                for x in items:
                    if x in _sum:
                        #_sum[x] += i[x]['value']
                        _lowhigh[x].append(makeZero(float('%.2f' % (i[x]['value'] if not i[x]['value'] == None else 0))))
                        _sum[x] += makeZero(float('%.2f' % (i[x]['value'] if not i[x]['value'] == None else 0)))
                    else:
                        _lowhigh[x] = [makeZero(float('%.2f' % (i[x]['value'] if not i[x]['value'] == None else 0)))]
                        _sum[x] = makeZero(float('%.2f' % (i[x]['value'] if not i[x]['value'] == None else 0)))
                if dtype == "firewall":
                    tmpu = []
                    tmpu = [makeZero('%.2f' % (i[k]['value'] if not i[k]['value'] == None else 0)) for k in items]
                    tmpu[0] = tmpu[1] + tmpu[2]
                    tmpu[3] = tmpu[4] + tmpu[5]
                    FinalData.append([_date] + tmpu)
                else:
                    FinalData.append([_date] + [makeZero('%.2f' % (i[k]['value'] if not i[k]['value'] == None else 0)) for k in items])
                # {'key_as_string': '2020-09-19T12:30:00.000Z', 'key': 1600518600000, 'doc_count': 125, 'CPUUsage': {'value': 5.435679975450039}, 'VCPU': {'value': 5.84}}
            sid = data['_scroll_id']
            scroll_size = len(data['hits']['hits'])
            break
        for x in list(_sum.keys()):
            _sum[x] = float('%.2f' % _sum[x])
        for x in _sum:
            _avg[x] = float('%.2f' % (_sum[x] / len(FinalData)))
        dTmp = {}
        for x in _lowhigh:
            dTmp[x] = {'low': min(_lowhigh[x]), 'high': max(_lowhigh[x])}
        FinalData.insert(0, ['DateTime'] + items)

        if dtype == "firewall":
            _sum['Traffic Total(speed)'] = _sum['Traffic In(speed)'] + _sum['Traffic Out(speed)']
            _sum['Traffic Total(volume)'] = _sum['Traffic In(volume)'] + _sum['Traffic Out(volume)']
            _avg['Traffic Total(speed)'] = _avg['Traffic In(speed)'] + _avg['Traffic Out(speed)']
            _avg['Traffic Total(volume)'] = _avg['Traffic In(volume)'] + _avg['Traffic Out(volume)']
            dTmp['Traffic Total(speed)']['low'] = dTmp['Traffic In(speed)']['low'] + dTmp['Traffic Out(speed)']['low']
            dTmp['Traffic Total(volume)']['low'] = dTmp['Traffic In(volume)']['low'] + dTmp['Traffic Out(volume)']['low']
            dTmp['Traffic Total(speed)']['high'] = dTmp['Traffic In(speed)']['high'] + dTmp['Traffic Out(speed)']['high']
            dTmp['Traffic Total(volume)']['high'] = dTmp['Traffic In(volume)']['high'] + dTmp['Traffic Out(volume)']['high']


        _2d = []
        # Low High
        lH = [i for i in dTmp]
        _2d = _2d + [['Low & High', '']]
        _2d.append([''] + [(x + 'MB') if x.__contains__('volume') else (x + _units) for x in lH])
        _2d.append(['Low'] + [dTmp[i]['low'] for i in lH])
        _2d.append(['High'] + [dTmp[i]['high'] for i in lH])
        _2d = _2d + [['', ''], ['', '']]
        # Average
        lH = [i for i in _avg]
        _2d = _2d + [['Average', '']]
        _2d.append([(x + 'MB') if x.__contains__('volume') else (x + _units) for x in lH])
        _2d.append([_avg[i] for i in lH])
        _2d = _2d + [['', ''], ['', '']]
        # Sum
        lH = [i for i in _sum]
        _2d = _2d + [['Sum', '']]
        _2d.append([(x + 'MB') if x.__contains__('volume') else (x + _units) for x in lH])
        _2d.append([_sum[i] for i in lH])
        _2d = _2d + [['', ''], ['', '']]
        # Data
        _2d = _2d + [['Data Sheet', '']]
        #FinalData[0][FinalData[0].index('Traffic Total(volume)Mbit/s')] = "Traffic Total(volume) MB"
        #FinalData[0][FinalData[0].index('Traffic In(volume)Mbit/s')] = "Traffic In(volume) MB"
        #FinalData[0][FinalData[0].index('Traffic Out(volume)Mbit/s')] = "Traffic Out(volume) MB"
        _2d = _2d + [[FinalData[0][0]] + [(x + 'MB') if x.__contains__('volume') else (x + _units) for x in FinalData[0][1:]]]
        _2d = _2d + FinalData[1:]
        out = export2XLSX(_2d)
        if out['result'] == 'success':
            return {'result': 'success', 'data': {'plots': FinalData, 'Sums': _sum, 'Average': _avg, 'LowHigh': dTmp, 'link': out['data']}}
        else:
            return {'result': 'success', 'data': {'plots': FinalData, 'Sums': _sum, 'Average': _avg, 'LowHigh': dTmp, 'link': ''}}
    except Exception as e:
        print(str(e))
        return {'result': 'failure', 'data': []}

def perfDataLatest(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:

                _category = dPayload['category']
                _item = dPayload['item']
                _name = dPayload['list']
                _metric = dPayload['metrics']
                timezone = dPayload["TimeZone"]  # "Asia/Kolkata"
                start_datetime = dPayload["start_datetime"]  # "2020-08-29 00:00"
                end_datetime = dPayload["end_datetime"]  # "2020-08-30 00:00"

                indexes = {'ESXi Host': 'ai-esxhost-perf-metrics',
                           'ESXi VM': 'ai-vm-perf-metrics',
                           'Datastore': 'ai-esxids-perf-metrics',
                           'KVM Host': 'ai-nghost-perf-metrics',
                           'KVM VM': 'ai-ngvm-perf-metrics',
                           'Firewall': 'ai-network-interface-bandwidth',
                           'Switch': 'ai-network-interface-switch-bandwidth'}

                mapper = {
                    'ESXi VM': {'CPU': {'select': ['CPUUsage'], 'meta': ['VCPU'], 'units': '%', 'plotter': ['CPUUsage']},
                               'Memory': {'select': ['MemoryUsage'], 'meta': ['TotalMemory', 'MemUnits'], 'units': 'GB', 'plotter': ['MemoryUsage']},
                               'Disk': {'select': [], 'meta': ['DiskCapacity', 'DiskUnits'], 'units': 'GB', 'plotter': []}, # Not Supported
                               'NIC': {'select': ['NetworkRx', 'NetworkTx'], 'meta': ['NetUnits'], 'units': 'Kbps', 'plotter': ['NetworkRx', 'NetworkTx']}},
                    'ESXi Host': {'CPU': {'select': ['CPU'], 'meta': [''], 'units': '%', 'plotter': ['CPU']},
                               'Memory': {'select': ['MEM'], 'meta': [''], 'units': '%', 'plotter': ['MEM']},
                               'Disk': {'select': ['DiskReadAverage', 'DiskWriteAverage'], 'meta': ['DiskUnits'], 'units': 'KBps', 'plotter': ['DiskReadAverage', 'DiskWriteAverage']},
                               'NIC': {'select': ['NetRx', 'NetTx'], 'meta': ['NetUnits'], 'units': 'KBps', 'plotter': ['NetRx', 'NetTx']}},
                    'Datastore': {'Disk': {'select': ['TotalSpace', 'FreeSpace'], 'meta': ['Units'], 'units': 'GB', 'plotter': ['TotalSpace', 'FreeSpace']}},
                    'KVM Host': {'CPU': {'select': ['CPU.Percent'], 'meta': ['CPU.Count'], 'units': '%', 'plotter': ['CPU.Percent']},
                               'Memory': {'select': ['Memory.Total', 'Memory.Free'], 'meta': [''], 'units': 'GiB', 'plotter': ['Memory.Total', 'Memory.Free']},
                               'Disk': {'select': ['DISK.Dynamic.Total', 'DISK.Dynamic.Free'], 'meta': ['', ''], 'units': 'GB', 'plotter': ['DISK.Dynamic.Total', 'DISK.Dynamic.Free']}, # Not Supported
                               'NIC': {'select': ['Interface.Dynamic.Bytes_recv', 'Interface.Dynamic.Bytes_sent'], 'meta': ['Units'], 'units': 'MiB/s',
                                       'plotter': ['Interface.Dynamic.Bytes_recv', 'Interface.Dynamic.Bytes_sent']}},
                    'KVM VM': {'CPU': {'select': ['cpu_percent'], 'meta': [], 'units': '%', 'plotter': ['CPU']},
                               'Memory':{'select': [], 'meta': [], 'units': '%', 'plotter': []}, # Not Supported
                               'Disk': {'select': ['disk.Dynamic.disk_percent'], 'meta': [], 'units': '%', 'plotter': ['Disk.Dynamic.percent']},
                               'NIC':{'select': ['interface.Dynamic.bytes_recv', 'interface.Dynamic.bytes_sent'], 'meta': [], 'units': 'MiB/s', 'plotter': ['Interface.Dynamic.bytes_recv', 'Interface.Dynamic.bytes_sent']}}, # Yet to confirm
                    'Firewall': {'NIC': {
                        'select': ['Traffic Total(speed)', 'Traffic In(speed)', 'Traffic Out(speed)', 'Traffic Total(volume)', 'Traffic In(volume)', 'Traffic Out(volume)'], 'units': 'Mbit/s',
                        'plotter': ['Traffic Total(speed)', 'Traffic In(speed)', 'Traffic Out(speed)']}},
                    'Switch': {'NIC': {
                        'select': ['Traffic Total(speed)', 'Traffic In(speed)', 'Traffic Out(speed)', 'Traffic Total(volume)', 'Traffic In(volume)', 'Traffic Out(volume)'], 'units': 'Mbit/s',
                        'plotter': ['Traffic Total(speed)', 'Traffic In(speed)', 'Traffic Out(speed)']}}
                }

                # DateTime Formatting
                user_tz = pytz.timezone(timezone)
                es_tz = pytz.timezone('GMT')
                user_s, user_e = user_tz.localize(dt.strptime(start_datetime, '%Y-%m-%d %H:%M')), user_tz.localize(
                    dt.strptime(end_datetime, '%Y-%m-%d %H:%M'))  # Declare I/P to particular timezone
                gmt_s, gmt_e = user_s.astimezone(es_tz), user_e.astimezone(es_tz)
                start_datetime = gmt_s.strftime('%Y-%m-%dT%H:%MZ')
                end_datetime = gmt_e.strftime('%Y-%m-%dT%H:%MZ')
                _bucket = math.ceil(((gmt_e.timestamp() - gmt_s.timestamp()) / 60) / 1000)

                _body = {'size': 1,
                         '_source': [],
                         'query': {'bool': {'filter': []}},
                         'aggs': {'Date': {'date_histogram': {'field': '@timestamp.GMT', 'fixed_interval': '{0}m'.format(_bucket)},
                                           'aggs': {}}}}

                if _category.lower().strip() == 'vcenter':
                    _hypip, _dcname, _object = _item.split('..')[0], _item.split('..')[1], _item.split('..')[2]
                    _index = indexes[_object]

                    if _object == 'ESXi VM':
                        _body['_source'] = ['@timestamp.GMT'] + mapper[_object][_metric]['select']
                        d, l = {}, []
                        l.append({'bool': {'should': [{'match_phrase': {'Hypervisor': _hypip}}]}})
                        l.append({'bool': {'should': [{'match_phrase': {'Name': _name}}]}})
                        d = {'bool': {'filter': l}}
                        _body['query']['bool']['filter'].append(d)
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Hypervisor': _hypip}})
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Name': _name}})
                        _body['query']['bool']['filter'].append({'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                        units = mapper[_object][_metric]['units']
                        plotters = mapper[_object][_metric]['plotter']
                        d = {}
                        for i in mapper[_object][_metric]['select']:
                            d[i] = {'avg': {'field': i}}
                        _body['aggs']['Date']['aggs'] = d
                        _out = esExec(_index, _body, es_tz, user_tz, _bucket, units)
                        if _out['result'] == 'success':
                            _out['units'] = units
                            _out['plotters'] = plotters
                            return json.dumps(_out)
                        else:
                            return json.dumps({'result': 'failure', 'data': 'no data'})
                    elif _object == 'ESXi Host':
                        _body['_source'] = ['@timestamp.GMT'] + mapper[_object][_metric]['select']
                        d, l = {}, []
                        l.append({'bool': {'should': [{'match_phrase': {'Hypervisor': _hypip}}]}})
                        l.append({'bool': {'should': [{'match_phrase': {'Name': _name}}]}})
                        d = {'bool': {'filter': l}}
                        _body['query']['bool']['filter'].append(d)
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Hypervisor': _hypip}})
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Name': _name}})
                        _body['query']['bool']['filter'].append(
                            {'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                        units = mapper[_object][_metric]['units']
                        plotters = mapper[_object][_metric]['plotter']
                        d = {}
                        for i in mapper[_object][_metric]['select']:
                            d[i] = {'avg': {'field': i}}
                        _body['aggs']['Date']['aggs'] = d
                        _out = esExec(_index, _body, es_tz, user_tz, _bucket, units)
                        if _out['result'] == 'success':
                            _out['units'] = units
                            _out['plotters'] = plotters
                            return json.dumps(_out)
                        else:
                            return json.dumps({'result': 'failure', 'data': 'no data'})
                    elif _object == 'Datastore':
                        _body['_source'] = ['@timestamp.GMT'] + mapper[_object][_metric]['select']
                        d, l = {}, []
                        l.append({'bool': {'should': [{'match_phrase': {'Hypervisor': _hypip}}]}})
                        l.append({'bool': {'should': [{'match_phrase': {'Name': _name}}]}})
                        d = {'bool': {'filter': l}}
                        _body['query']['bool']['filter'].append(d)
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Hypervisor': _hypip}})
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Name': _name}})
                        _body['query']['bool']['filter'].append(
                            {'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                        units = mapper[_object][_metric]['units']
                        plotters = mapper[_object][_metric]['plotter']
                        d = {}
                        for i in mapper[_object][_metric]['select']:
                            d[i] = {'avg': {'field': i}}
                        _body['aggs']['Date']['aggs'] = d
                        _out = esExec(_index, _body, es_tz, user_tz, _bucket, units)
                        if _out['result'] == 'success':
                            _out['units'] = units
                            _out['plotters'] = plotters
                            return json.dumps(_out)
                        else:
                            return json.dumps({'result': 'failure', 'data': 'no data'})

                elif _category.lower().strip() == 'kvm':
                    _index = indexes[_item]

                    if _item == 'KVM Host':
                        _body['_source'] = ['@timestamp.GMT'] + mapper[_item][_metric]['select']
                        d, l = {}, []
                        l.append({'bool': {'should': [{'match_phrase': {'IP': _name}}]}})
                        d = {'bool': {'filter': l}}
                        _body['query']['bool']['filter'].append(d)
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'IP': _name}})
                        _body['query']['bool']['filter'].append({'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                        units = mapper[_item][_metric]['units']
                        plotters = mapper[_item][_metric]['plotter']

                        if mapper[_item][_metric]['select'][0].lower().__contains__('.dynamic.'):
                            del _body['_source']
                            del _body['aggs']
                            _out = esExecWithAgg(_index, _body, es_tz, user_tz, _bucket)
                            if _out['result'] == 'success':
                                _lDyna = []
                                for x in mapper[_item][_metric]['select']:
                                    _lDyna += ["{0}.{1}.{2}".format(x.split('.')[0], i, x.split('.')[2]) for i in _out['data'][0][x.split('.')[0]]]

                                _body['_source'] = ['@timestamp.GMT'] + _lDyna
                                _body['aggs'] = {'Date': {'date_histogram': {'field': '@timestamp.GMT', 'fixed_interval': '{0}m'.format(_bucket)}, 'aggs': {}}}
                                d = {}
                                for i in _lDyna:
                                    d[i] = {'avg': {'field': i}}
                                _body['aggs']['Date']['aggs'] = d
                                _out = esExec(_index, _body, es_tz, user_tz, _bucket, units)
                                if _out['result'] == 'success':
                                    _out['units'] = units
                                    _out['plotters'] = _lDyna
                                    return json.dumps(_out)
                                else:
                                    return json.dumps({'result': 'failure', 'data': 'no data'})
                            else:
                                return json.dumps({'result': 'failure', 'data': 'no data'})

                        else:
                            d = {}
                            for i in mapper[_item][_metric]['select']:
                                d[i] = {'avg': {'field': i}}
                            _body['aggs']['Date']['aggs'] = d
                            print(_body)
                            _out = esExec(_index, _body, es_tz, user_tz, _bucket, units)
                            if _out['result'] == 'success':
                                _out['units'] = units
                                _out['plotters'] = plotters
                                return json.dumps(_out)
                            else:
                                return json.dumps({'result': 'failure', 'data': 'no data'})

                    elif _item == 'KVM VM':
                        _body['_source'] = ['@timestamp.GMT'] + mapper[_item][_metric]['select']
                        d, l = {}, []
                        l.append({'bool': {'should': [{'match_phrase': {'Name': _name}}]}})
                        d = {'bool': {'filter': l}}
                        _body['query']['bool']['filter'].append(d)
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Name': _name}})
                        _body['query']['bool']['filter'].append({'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                        units = mapper[_item][_metric]['units']
                        plotters = mapper[_item][_metric]['plotter']

                        if mapper[_item][_metric]['select'][0].lower().__contains__('.dynamic.'):
                            del _body['_source']
                            del _body['aggs']
                            _out = esExecWithAgg(_index, _body, es_tz, user_tz, _bucket)
                            if _out['result'] == 'success':
                                _lDyna = []
                                for x in mapper[_item][_metric]['select']:
                                    _lDyna += ["{0}.{1}.{2}".format(x.split('.')[0], i, x.split('.')[2]) for i in _out['data'][0][x.split('.')[0]]]

                                _body['_source'] = ['@timestamp.GMT'] + _lDyna
                                _body['aggs'] = {'Date': {'date_histogram': {'field': '@timestamp.GMT',
                                                                             'fixed_interval': '{0}m'.format(_bucket)},
                                                          'aggs': {}}}
                                d = {}
                                for i in _lDyna:
                                    d[i] = {'avg': {'field': i}}
                                _body['aggs']['Date']['aggs'] = d
                                _out = esExec(_index, _body, es_tz, user_tz, _bucket, units)
                                if _out['result'] == 'success':
                                    _out['units'] = units
                                    _out['plotters'] = _lDyna
                                    return json.dumps(_out)
                                else:
                                    return json.dumps({'result': 'failure', 'data': 'no data'})
                            else:
                                return json.dumps({'result': 'failure', 'data': 'no data'})

                        else:
                            d = {}
                            for i in mapper[_item][_metric]['select']:
                                d[i] = {'avg': {'field': i}}
                            _body['aggs']['Date']['aggs'] = d
                            print(_body)
                            _out = esExec(_index, _body, es_tz, user_tz, _bucket, units)
                            if _out['result'] == 'success':
                                _out['units'] = units
                                _out['plotters'] = plotters
                                return json.dumps(_out)
                            else:
                                return json.dumps({'result': 'failure', 'data': 'no data'})

                elif _category.lower().strip() == 'firewall':
                    _index = indexes['Firewall']
                    _body['_source'] = ['@timestamp.GMT'] + mapper[_category]['NIC']['select']
                    _body['query']['bool']['filter'] = []

                    d, l = {}, []
                    l.append({'bool': {'should': [{'match_phrase': {'Hostname': _name}}]}})
                    l.append({'bool': {'should': [{'match_phrase': {'Interface': _metric.split('..')[0]}}]}})
                    d = {'bool': {'filter': l}}
                    _body['query']['bool']['filter'].append(d)
                    #_body['query']['bool']['filter'].append({'match_phrase_prefix': {'Interface': _metric.split('..')[0]}})
                    # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Hostname': _name}})
                    # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Interface': _metric.split('..')[0]}})
                    _body['query']['bool']['filter'].append(
                        {'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                    units = mapper[_category]['NIC']['units']
                    plotters = mapper[_category]['NIC']['plotter']
                    d = {}
                    for i in mapper[_category]['NIC']['select']:
                        d[i] = {'sum': {'field': i}}
                    _body['aggs']['Date']['aggs'] = d
                    _out = esExec(_index, _body, es_tz, user_tz, _bucket, units, 'firewall')
                    if _out['result'] == 'success':
                        _out['units'] = units
                        _out['plotters'] = plotters
                        return json.dumps(_out)
                    else:
                        return json.dumps({'result': 'failure', 'data': 'no data'})

                elif _category.lower().strip() == 'switch':
                    _index = indexes['Switch']
                    _body['_source'] = ['@timestamp.GMT'] + mapper[_category]['NIC']['select']
                    d, l = {}, []
                    l.append({'bool': {'should': [{'match_phrase': {'Hostname': _name}}]}})
                    l.append({'bool': {'should': [{'match_phrase': {'Interface': _metric.split('..')[0]}}]}})
                    d = {'bool': {'filter': l}}
                    _body['query']['bool']['filter'].append(d)
                    # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Hostname': _name}})
                    # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Interface': _metric.split('..')[0]}})
                    _body['query']['bool']['filter'].append(
                        {'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                    units = mapper[_category]['NIC']['units']
                    plotters = mapper[_category]['NIC']['plotter']
                    d = {}
                    for i in mapper[_category]['NIC']['select']:
                        d[i] = {'sum': {'field': i}}
                    _body['aggs']['Date']['aggs'] = d
                    _out = esExec(_index, _body, es_tz, user_tz, _bucket, units)
                    if _out['result'] == 'success':
                        _out['units'] = units
                        _out['plotters'] = plotters
                        return json.dumps(_out)
                    else:
                        return json.dumps({'result': 'failure', 'data': 'no data'})

                return json.dumps({"result": "success", "data": ''})
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing



def esExec1(_pindex, _pbody, _es_tz, _user_tz, _bucket, _units, dtype = 'null'):
    try:
        es = Elasticsearch(dIPFQDN[location])
        result = es.search(index=_pindex, scroll='2m', body=_pbody, size=1000)
        FinalData, ft = [], True
        _sum, _avg = {}, {}
        _lowhigh = {}
        items = _pbody['_source'][1:]
        sid = result['_scroll_id']
        scroll_size = len(result['hits']['hits'])
        if scroll_size <= 0:
            return {"result": "failure", "data": "no data"}
        while scroll_size > 0:
            data = es.scroll(scroll_id=sid, scroll='{0}m'.format(_bucket))
            for i in result['aggregations']['Date']['buckets']:

                if dtype == "firewall" or dtype == "switch":
                    # Till 02-DEC-2020 23:39 All records are in MB Format
                    _ElasticDT = _es_tz.localize(dt.strptime(i['key_as_string'], '%Y-%m-%dT%H:%M:%S.%fZ'))
                    _bugdec2 = _es_tz.localize(dt.strptime('2020-12-01T18:29:59.000Z', '%Y-%m-%dT%H:%M:%S.%fZ'))
                    if _ElasticDT > _bugdec2:
                        for x in items:
                            i[x]['value'] = i[x]['value'] / (1024 * 1024)

                _date = _es_tz.localize(dt.strptime(i['key_as_string'], '%Y-%m-%dT%H:%M:%S.%fZ')).astimezone(_user_tz).strftime('%Y-%m-%d %H:%M')
                for x in items:
                    if x in _sum:
                        #_sum[x] += i[x]['value']
                        _lowhigh[x].append(makeZero(float('%.6f' % (i[x]['value'] if not i[x]['value'] == None else 0))))
                        _sum[x] += makeZero(float('%.6f' % (i[x]['value'] if not i[x]['value'] == None else 0)))
                    else:
                        _lowhigh[x] = [makeZero(float('%.6f' % (i[x]['value'] if not i[x]['value'] == None else 0)))]
                        _sum[x] = makeZero(float('%.6f' % (i[x]['value'] if not i[x]['value'] == None else 0)))
                if dtype == "firewall" or dtype == "switch":
                    tmpu = []
                    tmpu = [makeZero('%.6f' % (i[k]['value'] if not i[k]['value'] == None else 0)) for k in items]
                    tmpu[0] = tmpu[1] + tmpu[2]
                    tmpu[3] = tmpu[4] + tmpu[5]
                    FinalData.append([_date] + tmpu)
                else:
                    FinalData.append([_date] + [makeZero('%.6f' % (i[k]['value'] if not i[k]['value'] == None else 0)) for k in items])
                # {'key_as_string': '2020-09-19T12:30:00.000Z', 'key': 1600518600000, 'doc_count': 125, 'CPUUsage': {'value': 5.435679975450039}, 'VCPU': {'value': 5.84}}
            sid = data['_scroll_id']
            scroll_size = len(data['hits']['hits'])
            break
        for x in list(_sum.keys()):
            _sum[x] = float('%.6f' % _sum[x])
        for x in _sum:
            _avg[x] = float('%.6f' % (_sum[x] / len(FinalData)))
        dTmp = {}
        for x in _lowhigh:
            try:
                dTmp[x] = {'low': sorted(set(_lowhigh[x]))[1], 'high': max(_lowhigh[x])}
            except Exception as e:
                dTmp[x] = {'low': min(_lowhigh[x]), 'high': max(_lowhigh[x])}
        FinalData.insert(0, ['DateTime'] + items)

        if dtype == "firewall":
            # Compensate the total to in + out. This has to proper at source level
            _sum['Traffic Total(speed)'] = _sum['Traffic In(speed)'] + _sum['Traffic Out(speed)']
            _sum['Traffic Total(volume)'] = _sum['Traffic In(volume)'] + _sum['Traffic Out(volume)']
            _avg['Traffic Total(speed)'] = _avg['Traffic In(speed)'] + _avg['Traffic Out(speed)']
            _avg['Traffic Total(volume)'] = _avg['Traffic In(volume)'] + _avg['Traffic Out(volume)']
            dTmp['Traffic Total(speed)']['low'] = dTmp['Traffic In(speed)']['low'] + dTmp['Traffic Out(speed)']['low']
            dTmp['Traffic Total(volume)']['low'] = dTmp['Traffic In(volume)']['low'] + dTmp['Traffic Out(volume)']['low']
            dTmp['Traffic Total(speed)']['high'] = dTmp['Traffic In(speed)']['high'] + dTmp['Traffic Out(speed)']['high']
            dTmp['Traffic Total(volume)']['high'] = dTmp['Traffic In(volume)']['high'] + dTmp['Traffic Out(volume)']['high']

        # _2d = []
        # # Low High
        # lH = [i for i in dTmp]
        # _2d = _2d + [['Low & High', '']]
        # _2d.append([''] + [(x + 'MB') if x.__contains__('volume') else (x + _units) for x in lH])
        # _2d.append(['Low'] + [dTmp[i]['low'] for i in lH])
        # _2d.append(['High'] + [dTmp[i]['high'] for i in lH])
        # _2d = _2d + [['', ''], ['', '']]
        # # Average
        # lH = [i for i in _avg]
        # _2d = _2d + [['Average', '']]
        # _2d.append([(x + 'MB') if x.__contains__('volume') else (x + _units) for x in lH])
        # _2d.append([_avg[i] for i in lH])
        # _2d = _2d + [['', ''], ['', '']]
        # # Sum
        # lH = [i for i in _sum]
        # _2d = _2d + [['Sum', '']]
        # _2d.append([(x + 'MB') if x.__contains__('volume') else (x + _units) for x in lH])
        # _2d.append([_sum[i] for i in lH])
        # _2d = _2d + [['', ''], ['', '']]
        # # Data
        # _2d = _2d + [['Data Sheet', '']]
        # #FinalData[0][FinalData[0].index('Traffic Total(volume)Mbit/s')] = "Traffic Total(volume) MB"
        # #FinalData[0][FinalData[0].index('Traffic In(volume)Mbit/s')] = "Traffic In(volume) MB"
        # #FinalData[0][FinalData[0].index('Traffic Out(volume)Mbit/s')] = "Traffic Out(volume) MB"
        # _2d = _2d + [[FinalData[0][0]] + [(x + 'MB') if x.__contains__('volume') else (x + _units) for x in FinalData[0][1:]]]
        # _2d = _2d + FinalData[1:]
        # out = export2XLSX(_2d)
        # if out['result'] == 'success':
        #     return {'result': 'success', 'data': {'plots': FinalData, 'Sums': _sum, 'Average': _avg, 'LowHigh': dTmp, 'link': out['data']}}
        # else:
        #     return {'result': 'success', 'data': {'plots': FinalData, 'Sums': _sum, 'Average': _avg, 'LowHigh': dTmp, 'link': ''}}
        return {'result': 'success', 'data': {'plots': FinalData, 'Sums': _sum, 'Average': _avg, 'LowHigh': dTmp}}
    except Exception as e:
        print(str(e))
        return {'result': 'failure', 'data': []}

def generateXlS(js):
    try:
        _2d = []
        lh = js['data']['LowHigh']
        _2d.append(['Low & High', ''])
        _2d.append([''] + [i for i in lh])
        _2d.append(['Low'] + [lh[i]['low'] for i in lh])
        _2d.append(['High'] + [lh[i]['high'] for i in lh])
        _2d = _2d + [['', ''], ['', '']]
        del lh
        av = js['data']['Average']
        _2d.append(['Average', ''])
        _2d.append([i for i in av])
        _2d.append([av[i] for i in av])
        _2d = _2d + [['', ''], ['', '']]
        del av
        sm = js['data']['Sums']
        _2d.append(['Sum', ''])
        _2d.append([i for i in sm])
        _2d.append([sm[i] for i in sm])
        _2d = _2d + [['', ''], ['', '']]
        del sm
        gr = js['data']['grid']
        _2d.append(['Data Sheet', ''])
        _2d = _2d + gr
        del gr
        link = export2XLSX(_2d)
        js['data']['link'] = link['data']
        return js
    except Exception as e:
        js['data']['link'] = ''
        return js

def adjustUnits(js):
    try:

        # Grid
        adjustCopy = deepcopy(js['data']['plots'])
        speedIndex = [adjustCopy[0].index(i) for i in adjustCopy[0] if i.__contains__('speed')]
        for ind in range(1, len(adjustCopy)):
            tmp = adjustCopy[ind]
            for i in range(1, len(tmp)):
                if tmp[i] < 1 and tmp[i] > 0.000999:
                    tmp[i] = '{0} KBit/s'.format('%.3f' %(tmp[i] * 1024)) if i in speedIndex else '{0} KBytes'.format('%.3f' %(tmp[i] * 1024))
                elif tmp[i] <= 0.000999:
                    tmp[i] = '{0} Bit/s'.format(round(tmp[i] * 1024 * 1024)) if i in speedIndex else '{0} Bytes'.format(round(tmp[i] * 1024 * 1024))
                elif tmp[i] > 1024:
                    tmp[i] = '{0} GBit/s'.format('%.3f' % (tmp[i] / 1024)) if i in speedIndex else '{0} GBytes'.format('%.3f' % (tmp[i] / 1024))
                else:
                    tmp[i] = '{0} MBit/s'.format('%.3f' %(tmp[i])) if i in speedIndex else '{0} MBytes'.format('%.3f' %(tmp[i]))
            adjustCopy[ind] = tmp
            del tmp
        js['data']['grid'] = adjustCopy
        del adjustCopy

        # Plotting
        adjustCopy = deepcopy(js['data']['plots'])
        speedIndex = [adjustCopy[0].index(i) for i in adjustCopy[0] if i.__contains__('speed')]
        for ind in range(1, len(adjustCopy)):
            tmp = adjustCopy[ind]
            for i in range(1, len(tmp)):
                tmp[i] = '%.2f' % (tmp[i])
            adjustCopy[ind] = tmp
            del tmp
        js['data']['plots'] = adjustCopy
        del adjustCopy

        # Averages
        adjustCopy = deepcopy(js['data']['Average'])
        for eachK in adjustCopy:
            if adjustCopy[eachK] < 1 and adjustCopy[eachK] > 0.000999:
                adjustCopy[eachK] = '{0} KBit/s'.format('%.3f' %(adjustCopy[eachK] * 1024)) if eachK.__contains__(
                    'speed') else '{0} KBytes'.format('%.3f' %(adjustCopy[eachK] * 1024))
            elif adjustCopy[eachK] <= 0.000999:
                adjustCopy[eachK] = '{0} Bit/s'.format(round(adjustCopy[eachK] * 1024 * 1024)) if eachK.__contains__(
                    'speed') else '{0} Bytes'.format(round(adjustCopy[eachK] * 1024 * 1024))
            elif adjustCopy[eachK] > 1024:
                adjustCopy[eachK] = '{0} GBit/s'.format('%.3f' % (adjustCopy[eachK] / 1024)) if eachK.__contains__(
                    'speed') else '{0} GBytes'.format('%.3f' % (adjustCopy[eachK] / 1024))
            else:
                adjustCopy[eachK] = '{0} MBit/s'.format('%.3f' % (adjustCopy[eachK])) if eachK.__contains__(
                    'speed') else '{0} MBytes'.format('%.3f' % (adjustCopy[eachK]))
        js['data']['Average'] = adjustCopy
        del adjustCopy

        # Sum
        adjustCopy = deepcopy(js['data']['Sums'])
        for eachK in adjustCopy:
            if adjustCopy[eachK] < 1 and adjustCopy[eachK] > 0.000999:
                adjustCopy[eachK] = '{0} KBit/s'.format('%.3f' % (adjustCopy[eachK] * 1024)) if eachK.__contains__(
                    'speed') else '{0} KBytes'.format('%.3f' % (adjustCopy[eachK] * 1024))
            elif adjustCopy[eachK] <= 0.000999:
                adjustCopy[eachK] = '{0} Bit/s'.format(round(adjustCopy[eachK] * 1024 * 1024)) if eachK.__contains__(
                    'speed') else '{0} Bytes'.format(round(adjustCopy[eachK] * 1024 * 1024))
            elif adjustCopy[eachK] > 1024:
                adjustCopy[eachK] = '{0} GBit/s'.format('%.3f' % (adjustCopy[eachK] / 1024)) if eachK.__contains__(
                    'speed') else '{0} GBytes'.format('%.3f' % (adjustCopy[eachK] / 1024))
            else:
                adjustCopy[eachK] = '{0} MBit/s'.format('%.3f' % (adjustCopy[eachK])) if eachK.__contains__(
                    'speed') else '{0} MBytes'.format('%.3f' % (adjustCopy[eachK]))
        js['data']['Sums'] = adjustCopy
        del adjustCopy

        # Low and High
        adjustCopy = deepcopy(js['data']['LowHigh'])
        for eachK in adjustCopy:
            for lh in ['low', 'high']:
                if adjustCopy[eachK][lh] < 1 and adjustCopy[eachK][lh] > 0.000999:
                    adjustCopy[eachK][lh] = '{0} KBit/s'.format('%.3f' % (adjustCopy[eachK][lh] * 1024)) if eachK.__contains__(
                        'speed') else '{0} KBytes'.format('%.3f' % (adjustCopy[eachK][lh] * 1024))
                elif adjustCopy[eachK][lh] <= 0.000999:
                    adjustCopy[eachK][lh] = '{0} Bit/s'.format(
                        round(adjustCopy[eachK][lh] * 1024 * 1024)) if eachK.__contains__(
                        'speed') else '{0} Bytes'.format(round(adjustCopy[eachK][lh] * 1024 * 1024))
                elif adjustCopy[eachK][lh] > 1024:
                    adjustCopy[eachK][lh] = '{0} GBit/s'.format('%.3f' % (adjustCopy[eachK][lh] / 1024)) if eachK.__contains__(
                        'speed') else '{0} GBytes'.format('%.3f' % (adjustCopy[eachK][lh] / 1024))
                else:
                    adjustCopy[eachK][lh] = '{0} MBit/s'.format('%.3f' % (adjustCopy[eachK][lh])) if eachK.__contains__(
                        'speed') else '{0} MBytes'.format('%.3f' % (adjustCopy[eachK][lh]))
        js['data']['LowHigh'] = adjustCopy
        del adjustCopy

        return generateXlS(js)
    except Exception as e:
        js['data']['link'] = ''
        return js

def adjustUnitsOnlyFloating(js):
    try:

        # Grid
        adjustCopy = deepcopy(js['data']['plots'])
        speedIndex = [adjustCopy[0].index(i) for i in adjustCopy[0] if i.__contains__('speed')]
        for ind in range(1, len(adjustCopy)):
            tmp = adjustCopy[ind]
            for i in range(1, len(tmp)):
                tmp[i] = '%.2f' % (tmp[i])
            adjustCopy[ind] = tmp
            del tmp
        js['data']['grid'] = adjustCopy
        del adjustCopy

        # Plotting
        adjustCopy = deepcopy(js['data']['plots'])
        speedIndex = [adjustCopy[0].index(i) for i in adjustCopy[0] if i.__contains__('speed')]
        for ind in range(1, len(adjustCopy)):
            tmp = adjustCopy[ind]
            for i in range(1, len(tmp)):
                tmp[i] = '%.2f' % (tmp[i])
            adjustCopy[ind] = tmp
            del tmp
        js['data']['plots'] = adjustCopy
        del adjustCopy

        # Averages
        adjustCopy = deepcopy(js['data']['Average'])
        for eachK in adjustCopy:
            adjustCopy[eachK] = '%.2f' % (adjustCopy[eachK])
        js['data']['Average'] = adjustCopy
        del adjustCopy

        # Sum
        adjustCopy = deepcopy(js['data']['Sums'])
        for eachK in adjustCopy:
            adjustCopy[eachK] = '%.2f' % (adjustCopy[eachK])
        js['data']['Sums'] = adjustCopy
        del adjustCopy

        # Low and High
        adjustCopy = deepcopy(js['data']['LowHigh'])
        for eachK in adjustCopy:
            for lh in ['low', 'high']:
                adjustCopy[eachK][lh] = '%.2f' % (adjustCopy[eachK][lh])
        js['data']['LowHigh'] = adjustCopy
        del adjustCopy

        return generateXlS(js)
    except Exception as e:
        js['data']['link'] = ''
        return js

def perfDataLatest1(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:

                _category = dPayload['category']
                _item = dPayload['item']
                _name = dPayload['list']
                _metric = dPayload['metrics']
                timezone = dPayload["TimeZone"]  # "Asia/Kolkata"
                start_datetime = dPayload["start_datetime"]  # "2020-08-29 00:00"
                end_datetime = dPayload["end_datetime"]  # "2020-08-30 00:00"

                indexes = {'ESXi Host': 'ai-esxhost-perf-metrics',
                           'ESXi VM': 'ai-vm-perf-metrics',
                           'Datastore': 'ai-esxids-perf-metrics',
                           'KVM Host': 'ai-nghost-perf-metrics',
                           'KVM VM': 'ai-ngvm-perf-metrics',
                           'Firewall': 'ai-network-interface-bandwidth',
                           'Switch': 'ai-network-interface-switch-bandwidth',
                           'Load Balancer': _item}

                mapper = {
                    'ESXi VM': {'CPU': {'select': ['CPUUsage'], 'meta': ['VCPU'], 'units': '%', 'plotter': ['CPUUsage']},
                               'Memory': {'select': ['MemoryUsage'], 'meta': ['TotalMemory', 'MemUnits'], 'units': '%', 'plotter': ['MemoryUsage']},
                               'Disk': {'select': [], 'meta': ['DiskCapacity', 'DiskUnits'], 'units': 'GB', 'plotter': []}, # Not Supported
                               'NIC': {'select': ['NetworkRx', 'NetworkTx'], 'meta': ['NetUnits'], 'units': 'Kbps', 'plotter': ['NetworkRx', 'NetworkTx']}},
                    'ESXi Host': {'CPU': {'select': ['CPU'], 'meta': [''], 'units': '%', 'plotter': ['CPU']},
                               'Memory': {'select': ['MEM'], 'meta': [''], 'units': '%', 'plotter': ['MEM']},
                               'Disk': {'select': ['DiskReadAverage', 'DiskWriteAverage'], 'meta': ['DiskUnits'], 'units': 'KBps', 'plotter': ['DiskReadAverage', 'DiskWriteAverage']},
                               'NIC': {'select': ['NetRx', 'NetTx'], 'meta': ['NetUnits'], 'units': 'KBps', 'plotter': ['NetRx', 'NetTx']}},
                    'Datastore': {'Disk': {'select': ['TotalSpace', 'FreeSpace'], 'meta': ['Units'], 'units': 'GB', 'plotter': ['TotalSpace', 'FreeSpace']}},
                    'KVM Host': {'CPU': {'select': ['CPU.Percent'], 'meta': ['CPU.Count'], 'units': '%', 'plotter': ['CPU.Percent']},
                               'Memory': {'select': ['Memory.Total', 'Memory.Free'], 'meta': [''], 'units': 'GB', 'plotter': ['Memory.Total', 'Memory.Free']},
                               'Disk': {'select': ['DISK.Dynamic.Total', 'DISK.Dynamic.Free'], 'meta': ['', ''], 'units': 'GB', 'plotter': ['DISK.Dynamic.Total', 'DISK.Dynamic.Free']}, # Not Supported
                               'NIC': {'select': ['Interface.Dynamic.Bytes_recv', 'Interface.Dynamic.Bytes_sent'], 'meta': ['Units'], 'units': 'MB/s',
                                       'plotter': ['Interface.Dynamic.Bytes_recv', 'Interface.Dynamic.Bytes_sent']}},
                    'KVM VM': {'CPU': {'select': ['cpu_percent'], 'meta': [], 'units': '%', 'plotter': ['cpu_percent']},
                               'Memory':{'select': [], 'meta': [], 'units': '%', 'plotter': []}, # Not Supported
                               'Disk': {'select': ['disk.Dynamic.disk_percent'], 'meta': [], 'units': '%', 'plotter': ['Disk.Dynamic.percent']},
                               'NIC':{'select': ['interface.Dynamic.bytes_recv', 'interface.Dynamic.bytes_sent'], 'meta': [], 'units': 'MB/s', 'plotter': ['Interface.Dynamic.bytes_recv', 'Interface.Dynamic.bytes_sent']}}, # Yet to confirm
                    'Firewall': {'NIC': {
                        'select': ['Traffic Total(speed)', 'Traffic In(speed)', 'Traffic Out(speed)', 'Traffic Total(volume)', 'Traffic In(volume)', 'Traffic Out(volume)'], 'units': 'Mbit/s',
                        'plotter': ['Traffic Total(speed)', 'Traffic In(speed)', 'Traffic Out(speed)']}},
                    'Switch': {'NIC': {
                        'select': ['Traffic Total(speed)', 'Traffic In(speed)', 'Traffic Out(speed)', 'Traffic Total(volume)', 'Traffic In(volume)', 'Traffic Out(volume)'], 'units': 'Mbit/s',
                        'plotter': ['Traffic Total(speed)', 'Traffic In(speed)', 'Traffic Out(speed)']}},
                    'Load Balancer': {
                        'Connection_Count': {'select': ['report_data.current_connection_count'], 'meta': [], 'units': '', 'plotters': ['report_data.current_connection_count']},
                        'Outstanding_Request': {'select': ['report_data.outstanding_request'], 'meta': [], 'units': '', 'plotters': ['report_data.outstanding_request']},
                        'Bytes_IN': {'select': ['report_data.bytes_in'], 'meta': [], 'units': 'Bytes', 'plotters': ['report_data.bytes_in']},
                        'Bytes_OUT': {'select': ['report_data.bytes_out'], 'meta': [], 'units': 'Bytes', 'plotters': ['report_data.bytes_out']},
                        'Packets_IN': {'select': ['report_data.packets_in'], 'meta': [], 'units': '', 'plotters': ['report_data.packets_in']},
                        'Packets_OUT': {'select': ['report_data.packets_out'], 'meta': [], 'units': '', 'plotters': ['report_data.packets_out']},
                        'Average_Bandwidth_IN': {'select': ['report_data.average_bandwidth_in'], 'meta': [], 'units': '', 'plotters': ['report_data.average_bandwidth_in']},
                        'Average_Bandwidth_Out': {'select': ['report_data.average_bandwidth_out'], 'meta': [], 'units': '', 'plotters': ['report_data.average_bandwidth_out']}
                    }
                }

                # DateTime Formatting
                user_tz = pytz.timezone(timezone)
                es_tz = pytz.timezone('GMT')
                user_s, user_e = user_tz.localize(dt.strptime(start_datetime, '%Y-%m-%d %H:%M')), user_tz.localize(
                    dt.strptime(end_datetime, '%Y-%m-%d %H:%M'))  # Declare I/P to particular timezone
                gmt_s, gmt_e = user_s.astimezone(es_tz), user_e.astimezone(es_tz)

                start_datetime = gmt_s.strftime('%Y-%m-%dT%H:%MZ')
                end_datetime = gmt_e.strftime('%Y-%m-%dT%H:%MZ')

                _bucket = math.ceil(((gmt_e.timestamp() - gmt_s.timestamp()) / 60) / 1440)

                _body = {'size': 1,
                         '_source': [],
                         'query': {'bool': {'filter': []}},
                         'aggs': {'Date': {'date_histogram': {'field': '@timestamp.GMT', 'fixed_interval': '{0}m'.format(_bucket)},
                                           'aggs': {}}}}

                if _category.lower().strip() == 'vcenter':
                    _hypip, _dcname, _object = _item.split('..')[0], _item.split('..')[1], _item.split('..')[2]
                    _index = indexes[_object]

                    if _object == 'ESXi VM':
                        _body['_source'] = ['@timestamp.GMT'] + mapper[_object][_metric]['select']
                        d, l = {}, []
                        l.append({'bool': {'should': [{'match_phrase': {'Hypervisor': _hypip}}]}})
                        l.append({'bool': {'should': [{'match_phrase': {'Name': _name}}]}})
                        d = {'bool': {'filter': l}}
                        _body['query']['bool']['filter'].append(d)
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Hypervisor': _hypip}})
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Name': _name}})
                        _body['query']['bool']['filter'].append({'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                        units = mapper[_object][_metric]['units']
                        plotters = mapper[_object][_metric]['plotter']
                        d = {}
                        for i in mapper[_object][_metric]['select']:
                            d[i] = {'sum': {'field': i}}
                        _body['aggs']['Date']['aggs'] = d
                        _out = esExec1(_index, _body, es_tz, user_tz, _bucket, units)
                        if _out['result'] == 'success':
                            _out['units'] = units
                            _out['plotters'] = plotters
                            return json.dumps(adjustUnitsOnlyFloating(_out))
                        else:
                            return json.dumps({'result': 'failure', 'data': 'no data'})
                    elif _object == 'ESXi Host':
                        _body['_source'] = ['@timestamp.GMT'] + mapper[_object][_metric]['select']
                        d, l = {}, []
                        l.append({'bool': {'should': [{'match_phrase': {'Hypervisor': _hypip}}]}})
                        l.append({'bool': {'should': [{'match_phrase': {'Name': _name}}]}})
                        d = {'bool': {'filter': l}}
                        _body['query']['bool']['filter'].append(d)
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Hypervisor': _hypip}})
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Name': _name}})
                        _body['query']['bool']['filter'].append(
                            {'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                        units = mapper[_object][_metric]['units']
                        plotters = mapper[_object][_metric]['plotter']
                        d = {}
                        for i in mapper[_object][_metric]['select']:
                            d[i] = {'avg': {'field': i}}
                        _body['aggs']['Date']['aggs'] = d
                        _out = esExec1(_index, _body, es_tz, user_tz, _bucket, units)
                        if _out['result'] == 'success':
                            _out['units'] = units
                            _out['plotters'] = plotters
                            return json.dumps(adjustUnitsOnlyFloating(_out))
                        else:
                            return json.dumps({'result': 'failure', 'data': 'no data'})
                    elif _object == 'Datastore':
                        _body['_source'] = ['@timestamp.GMT'] + mapper[_object][_metric]['select']
                        d, l = {}, []
                        l.append({'bool': {'should': [{'match_phrase': {'Hypervisor': _hypip}}]}})
                        l.append({'bool': {'should': [{'match_phrase': {'Name': _name}}]}})
                        d = {'bool': {'filter': l}}
                        _body['query']['bool']['filter'].append(d)
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Hypervisor': _hypip}})
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Name': _name}})
                        _body['query']['bool']['filter'].append(
                            {'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                        units = mapper[_object][_metric]['units']
                        plotters = mapper[_object][_metric]['plotter']
                        d = {}
                        for i in mapper[_object][_metric]['select']:
                            d[i] = {'avg': {'field': i}}
                        _body['aggs']['Date']['aggs'] = d
                        _out = esExec1(_index, _body, es_tz, user_tz, _bucket, units)
                        if _out['result'] == 'success':
                            _out['units'] = units
                            _out['plotters'] = plotters
                            return json.dumps(adjustUnitsOnlyFloating(_out))
                        else:
                            return json.dumps({'result': 'failure', 'data': 'no data'})

                elif _category.lower().strip() == 'kvm':
                    _index = indexes[_item]

                    if _item == 'KVM Host':
                        _body['_source'] = ['@timestamp.GMT'] + mapper[_item][_metric]['select']
                        d, l = {}, []
                        l.append({'bool': {'should': [{'match_phrase': {'IP': _name}}]}})
                        d = {'bool': {'filter': l}}
                        _body['query']['bool']['filter'].append(d)
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'IP': _name}})
                        _body['query']['bool']['filter'].append({'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                        units = mapper[_item][_metric]['units']
                        plotters = mapper[_item][_metric]['plotter']

                        if mapper[_item][_metric]['select'][0].lower().__contains__('.dynamic.'):
                            del _body['_source']
                            del _body['aggs']
                            _out = esExecWithAgg(_index, _body, es_tz, user_tz, _bucket)
                            if _out['result'] == 'success':
                                _lDyna = []
                                for x in mapper[_item][_metric]['select']:
                                    _lDyna += ["{0}.{1}.{2}".format(x.split('.')[0], i, x.split('.')[2]) for i in _out['data'][0][x.split('.')[0]]]

                                _body['_source'] = ['@timestamp.GMT'] + _lDyna
                                _body['aggs'] = {'Date': {'date_histogram': {'field': '@timestamp.GMT', 'fixed_interval': '{0}m'.format(_bucket)}, 'aggs': {}}}
                                d = {}
                                for i in _lDyna:
                                    d[i] = {'avg': {'field': i}}
                                _body['aggs']['Date']['aggs'] = d
                                _out = esExec1(_index, _body, es_tz, user_tz, _bucket, units)
                                if _out['result'] == 'success':
                                    _out['units'] = units
                                    _out['plotters'] = _lDyna
                                    return json.dumps(adjustUnitsOnlyFloating(_out))
                                else:
                                    return json.dumps({'result': 'failure', 'data': 'no data'})
                            else:
                                return json.dumps({'result': 'failure', 'data': 'no data'})

                        else:
                            d = {}
                            for i in mapper[_item][_metric]['select']:
                                d[i] = {'avg': {'field': i}}
                            _body['aggs']['Date']['aggs'] = d
                            print(_body)
                            _out = esExec1(_index, _body, es_tz, user_tz, _bucket, units)
                            if _out['result'] == 'success':
                                _out['units'] = units
                                _out['plotters'] = plotters
                                return json.dumps(adjustUnitsOnlyFloating(_out))
                            else:
                                return json.dumps({'result': 'failure', 'data': 'no data'})

                    elif _item == 'KVM VM':
                        _body['_source'] = ['@timestamp.GMT'] + mapper[_item][_metric]['select']
                        d, l = {}, []
                        l.append({'bool': {'should': [{'match_phrase': {'Name': _name}}]}})
                        d = {'bool': {'filter': l}}
                        _body['query']['bool']['filter'].append(d)
                        # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Name': _name}})
                        _body['query']['bool']['filter'].append({'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                        units = mapper[_item][_metric]['units']
                        plotters = mapper[_item][_metric]['plotter']

                        if mapper[_item][_metric]['select'][0].lower().__contains__('.dynamic.'):
                            del _body['_source']
                            del _body['aggs']
                            _out = esExecWithAgg(_index, _body, es_tz, user_tz, _bucket)
                            if _out['result'] == 'success':
                                _lDyna = []
                                for x in mapper[_item][_metric]['select']:
                                    _lDyna += ["{0}.{1}.{2}".format(x.split('.')[0], i, x.split('.')[2]) for i in _out['data'][0][x.split('.')[0]]]

                                _body['_source'] = ['@timestamp.GMT'] + _lDyna
                                _body['aggs'] = {'Date': {'date_histogram': {'field': '@timestamp.GMT',
                                                                             'fixed_interval': '{0}m'.format(_bucket)},
                                                          'aggs': {}}}
                                d = {}
                                for i in _lDyna:
                                    d[i] = {'avg': {'field': i}}
                                _body['aggs']['Date']['aggs'] = d
                                _out = esExec1(_index, _body, es_tz, user_tz, _bucket, units)
                                if _out['result'] == 'success':
                                    _out['units'] = units
                                    _out['plotters'] = _lDyna
                                    return json.dumps(adjustUnitsOnlyFloating(_out))
                                else:
                                    return json.dumps({'result': 'failure', 'data': 'no data'})
                            else:
                                return json.dumps({'result': 'failure', 'data': 'no data'})

                        else:
                            d = {}
                            for i in mapper[_item][_metric]['select']:
                                d[i] = {'avg': {'field': i}}
                            _body['aggs']['Date']['aggs'] = d
                            print(_body)
                            _out = esExec1(_index, _body, es_tz, user_tz, _bucket, units)
                            if _out['result'] == 'success':
                                _out['units'] = units
                                _out['plotters'] = plotters
                                return json.dumps(adjustUnitsOnlyFloating(_out))
                            else:
                                return json.dumps({'result': 'failure', 'data': 'no data'})

                elif _category.lower().strip() == 'firewall':
                    _index = indexes['Firewall']
                    _body['_source'] = ['@timestamp.GMT'] + mapper[_category]['NIC']['select']
                    _body['query']['bool']['filter'] = []

                    d, l = {}, []
                    l.append({'bool': {'should': [{'match_phrase': {'Hostname': _name}}]}})
                    l.append({'bool': {'should': [{'match_phrase': {'Interface': _metric.split('..')[0]}}]}})
                    d = {'bool': {'filter': l}}
                    _body['query']['bool']['filter'].append(d)
                    #_body['query']['bool']['filter'].append({'match_phrase_prefix': {'Interface': _metric.split('..')[0]}})
                    # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Hostname': _name}})
                    # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Interface': _metric.split('..')[0]}})
                    _body['query']['bool']['filter'].append(
                        {'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                    units = mapper[_category]['NIC']['units']
                    plotters = mapper[_category]['NIC']['plotter']
                    d = {}
                    for i in mapper[_category]['NIC']['select']:
                        d[i] = {'sum': {'field': i}}
                    _body['aggs']['Date']['aggs'] = d
                    _out = esExec1(_index, _body, es_tz, user_tz, _bucket, units, 'firewall')
                    if _out['result'] == 'success':
                        _out['units'] = units
                        _out['plotters'] = plotters
                        return json.dumps(adjustUnits(_out))
                    else:
                        return json.dumps({'result': 'failure', 'data': 'no data'})

                elif _category.lower().strip() == 'switch':
                    _index = indexes['Switch']
                    _body['_source'] = ['@timestamp.GMT'] + mapper[_category]['NIC']['select']
                    d, l = {}, []
                    l.append({'bool': {'should': [{'match_phrase': {'Hostname': _name}}]}})
                    l.append({'bool': {'should': [{'match_phrase': {'Interface': _metric.split('..')[0]}}]}})
                    d = {'bool': {'filter': l}}
                    _body['query']['bool']['filter'].append(d)
                    # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Hostname': _name}})
                    # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Interface': _metric.split('..')[0]}})
                    _body['query']['bool']['filter'].append(
                        {'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                    units = mapper[_category]['NIC']['units']
                    plotters = mapper[_category]['NIC']['plotter']
                    d = {}
                    for i in mapper[_category]['NIC']['select']:
                        d[i] = {'sum': {'field': i}}
                    _body['aggs']['Date']['aggs'] = d
                    _out = esExec1(_index, _body, es_tz, user_tz, _bucket, units, 'switch')
                    if _out['result'] == 'success':
                        _out['units'] = units
                        _out['plotters'] = plotters
                        return json.dumps(adjustUnits(_out))
                    else:
                        return json.dumps({'result': 'failure', 'data': 'no data'})

                elif _category.lower().strip() == 'load balancer':
                    _index = indexes['Load Balancer']
                    _body['_source'] = ['@timestamp.GMT'] + mapper[_category][_metric]['select']
                    d, l = {}, []
                    l.append({'bool': {'should': [{'match_phrase': {'instance_id': _name}}]}})
                    l.append({'bool': {'should': [{'match_phrase': {'service_name': _name}}]}})
                    d = {'bool': {'filter': l}}
                    _body['query']['bool']['filter'].append(d)
                    # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Hostname': _name}})
                    # _body['query']['bool']['filter'].append({'match_phrase_prefix': {'Interface': _metric.split('..')[0]}})
                    _body['query']['bool']['filter'].append(
                        {'range': {'@timestamp.GMT': {'gte': start_datetime, 'lte': end_datetime}}})
                    units = mapper[_category][_metric]['units']
                    plotters = mapper[_category][_metric]['plotters']
                    d = {}
                    for i in mapper[_category][_metric]['select']:
                        d[i] = {'sum': {'field': i}}
                    _body['aggs']['Date']['aggs'] = d
                    _out = esExec1(_index, _body, es_tz, user_tz, _bucket, units)
                    if _out['result'] == 'success':
                        _out['units'] = units
                        _out['plotters'] = plotters
                        if units.lower() == "bytes":
                            return json.dumps(adjustUnits(_out))
                        else:
                            return json.dumps(adjustUnitsOnlyFloating(_out))
                    else:
                        return json.dumps({'result': 'failure', 'data': 'no data'})


                return json.dumps({"result": "success", "data": 'no data'})
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing
