#============================================================================================
# Author        : Dinesh Balaraman
# Created On    :
# (c) Copyright by Autointelli Technologies
#============================================================================================

from flask import Flask, jsonify, request
import json
from services.utils import ConnPostgreSQL as conn
from services.utils import sessionkeygen as sess
import nmap
import services.utils.ConnWebSocket as ws
import services.utils.ED_AES256 as aes
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from datetime import datetime
from services.utils.ConnLog import create_log_file
from services.utils.ValidatorSession import chkValidRequest, chkKeyExistsInHeader, lam_api_key_invalid, lam_api_key_missing
import services.utils.LFColors as lfc
import services.utils.ConnMQ as connmq
import services.utils.validator_many as val
from datetime import datetime as dt
from services.utils import ConnArcon as pam1

lfcObj = lfc.bcolors()
CERROR, CWARN = lfcObj.printerr, lfcObj.printwar

logObj = create_log_file()
if not logObj:
    CERROR("Not able to create logfile")
    exit(0)
logERROR, logWARN, logINFO = logObj.error, logObj.warn, logObj.info

lam_api_key_missing = lam_api_key_missing()
lam_api_key_invalid = lam_api_key_invalid()
location, pam = "", ""


try:
    location = json.load(open('/etc/autointelli/autointelli.conf', 'r'))['audit_video']
    pam = json.load(open('/etc/autointelli/autointelli.conf', 'r'))['pvault']
except Exception as e:
    print("Couldn't start the CRON because the location is missing in conf file")

def logAndRet(s, x):
    if s == "failure":
        logERROR(x)
    else:
        logINFO(x)
    return json.dumps({"result": s, "data": x})

def initiateRESTDiscovery(dPayload):
    try:
        sQuery = "select is_on from is_discovery_on"
        dRet = conn.returnSelectQueryResult(sQuery)
        if dRet['result'] == 'success':
            flag = dRet['data'][0]['is_on']
            if flag == 'N':
                dRet = connmq.send2MQ(pQueue='discovery_socket', pExchange='automation', pRoutingKey='discovery_socket',
                                      pData=json.dumps(dPayload))
                iQuery = "update is_discovery_on set is_on='Y'"
                dRet = conn.returnInsertResult(iQuery)
                if dRet['result'] == 'success':
                    return json.dumps({'result': 'success', 'data': 'Discovery initiated successfully.'})
                else:
                    return json.dumps({'result': 'failure', 'data': 'Failed to initiate. Try after sometime.'})
            else:
                return json.dumps({'result': 'failure', 'data': 'Discovery Initiation is already in progress.'})
        else:
            return json.dumps({'result': 'failure', 'data': 'Failed to initiate. Try after sometime.'})
    except Exception as e:
        return json.dumps({'result': 'failure', 'data': 'Failed to initiate. Try after sometime.'})

def callback_result(host, scan_result):
    print('-' * 50)
    sIPAddress, sOS = "", ""
    sIPAddress = host
    if scan_result['nmap']['scanstats']['uphosts'] == '1':
        sOS = scan_result['scan'][host]['osmatch'][0]['osclass'][0]['osfamily']
        # Don't discover if the machine is already available
        sChkQuery = "select count(machine_id) total from ai_machine where ip_address='%s'" % sIPAddress
        dRet = conn.returnSelectQueryResult(sChkQuery)
        if dRet['result'] == 'success':
            if dRet['data'][0]['total'] == 0:
                # Before insert, remove data from discovery table
                sDelQuery = "delete from ai_device_discovery where ip_address='" + sIPAddress.strip() + "'"
                dRet = conn.returnInsertResult(sDelQuery)
                # Insert data
                sDisQuery = "insert into ai_device_discovery(ip_address, operating_system, gf_yn) values('%s', '%s', 'N')" % (
                    sIPAddress, sOS)
                dRet = conn.returnInsertResult(sDisQuery)
                if dRet["result"] == "success":
                    pass
                    # Redirect the result to socket
                    # dRet = ws.pushToWebSocket(pModule='discovery', pInfoType='json',
                    #                           pData={"ipaddress": sIPAddress, "OS": sOS}, pAction='create')
                    # print(dRet)
                else:
                    logERROR("Failed to insert data into database: {0}".format(dRet["data"]))
            else:
                logINFO("{0} already available. So, discovery is skipped".format(host))
        else:
            logERROR("Couldn't fetch any information from database: {0}".format(dRet["data"]))
    else:
        logERROR("{0} is not reachable".format(host))

def startDiscovery(dPayload):
    """pIPRange: IP Range => '192.168.1.100-110' """
    # Discovery Initiation Notification
    logINFO("Discovery Started")
    # dRet = ws.pushToWebSocket(pModule='discovery', pInfoType='msg', pData='Discovery Initiated', pAction='notify')
    # Begin Scanning
    nma = nmap.PortScannerAsync()
    pIPRange = dPayload["ip_range"]
    if pIPRange.__contains__('-'):
        l = pIPRange.split('-')
        _3Octet = '.'.join(l[0].split('.')[:-1])
        for eachIP in range(int(l[0].split('.')[-1]), int(l[1]) + 1):
            ip = _3Octet + '.' + str(eachIP)
            logINFO("SCAN:{0}".format(ip + ('.' * 10) + pIPRange))
            nma.scan(hosts=ip, arguments='-O -p 22-443', callback=callback_result)
            while nma.still_scanning():
                print('.' * 3)
                nma.wait(1)
    else:
        nma.scan(hosts=pIPRange, arguments='-O -p 22-443', callback=callback_result)
        while nma.still_scanning():
            print('.' * 3)
            nma.wait(1)
    #nma.scan(hosts=pIPRange, arguments='-O -p 22-443', callback=callback_result)
    #while nma.still_scanning():
    #    nma.wait(5)

    # Discovery Completion Notification
    # dRet = ws.pushToWebSocket(pModule='discovery', pInfoType='msg', pData='Discovery Completed', pAction='notify')
    logINFO("Discovery Completed")
    iQuery = "update is_discovery_on set is_on='N'"
    dRet = conn.returnInsertResult(iQuery)
    return {'result': 'success'}

# Listing Devices from Table
def getDeviceYetToDiscoverList():
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = "select pk_discovery_id discovery_id, ip_address, operating_system, cred_name, cred_id, username, password, sudo_yn from ai_device_discovery left join ai_device_credentials on(fk_cred_id=cred_id) where gf_yn='N' "
                dRet = conn.returnSelectQueryResult(sQuery)
                isQuery = "select is_on from is_discovery_on"
                isRet = conn.returnSelectQueryResult(isQuery)
                if dRet["result"] == "success" and isRet['result'] == 'success':
                    final = {'grid': dRet['data'], 'flag': isRet['data'][0]['is_on']}
                    return json.dumps({'result': 'success', 'data': final})
                else:
                    logERROR("Unable to fetch Discovery List from database.")
                    return json.dumps(dRet.update({"error": "Unable to fetch Discovery List from database."}))
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

# Listing Devices from Table
def getDeviceYetToDiscoverList4machine():
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = "select pk_discovery_id discovery_id, ip_address, operating_system, cred_name, cred_id, cred_type, username, password, sudo_yn, port from ai_device_discovery left join ai_device_credentials on fk_cred_id=cred_id and cred_type=cred_type where gf_yn='N' and active_yn='Y';"
                dRet = conn.returnSelectQueryResult(sQuery)
                if dRet["result"] == "success":
                    dFinal = []
                    k = '@ut0!ntell!'.encode()
                    for i in dRet["data"]:
                        passwd = i["password"].encode()
                        i["password"] = aes.decrypt(passwd, k).decode('utf-8')
                        dFinal.append(i)
                    return json.dumps({"result":"success", "data": dFinal})
                else:
                    logERROR("Unable to fetch Discovery List from database.")
                    return json.dumps(dRet.update({"error": "Unable to fetch Discovery List from database."}))
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

# Attach Credentials with Machine
def createMappingDiscoveryAndCredentials(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                if "mappers" in dPayload.keys():
                    s, f = 0, 0
                    for eachItem in dPayload["mappers"]:
                        iUpdate = "update ai_device_discovery set fk_cred_id=(select cred_id from ai_device_credentials where lower(cred_name)='" + eachItem[1].lower() + "') where ip_address='" + eachItem[0] + "'"
                        # iUpdate = "update ai_device_discovery set fk_cred_id={0} where ip_address='{1}'".format(eachItem[1], eachItem[0])
                        dRet = conn.returnInsertResult(iUpdate)
                        if dRet["result"] == "success":
                            s += 1
                            sQuery = "select ip_address, operating_system, fk_cred_id from ai_device_discovery where ip_address='{0}'".format(eachItem[0])
                            sRet = conn.returnSelectQueryResult(sQuery)
                            if sRet['result'] == 'success':
                                del sRet['result']
                                mRet = connmq.send2MQ(pQueue='d2m', pExchange='automationengine', pRoutingKey='d2m', pData=json.dumps(sRet))
                                print(mRet)
                        else:
                            f += 1
                    return logAndRet("success", "Successfully Mapped Credentials")
                else:
                    return logAndRet("failure", "Invalid Payload")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def createMappingDiscoveryAndCredentialsOld(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                if "mappers" in dPayload.keys():
                    s, f = 0, 0
                    for eachItem in dPayload["mappers"]:
                        iUpdate = "update ai_device_discovery set fk_cred_id=(select cred_id from ai_device_credentials where lower(cred_name)='" + eachItem[1].lower() + "') where ip_address='" + eachItem[0] + "'"
                        dRet = conn.returnInsertResult(iUpdate)
                        if dRet["result"] == "success":
                            s += 1
                        else:
                            f += 1
                    return logAndRet("success", "Successfully Mapped Credentials")
                else:
                    return logAndRet("failure", "Invalid Payload")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

# De-Attach Credentials with Machine
def deAttachCredentailsFromMachine(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                if "device_list" in dPayload.keys():
                    iUpdate = "update ai_device_discovery set fk_cred_id=null where ip_address in('" + "','".join(dPayload["device_list"]) + "')"
                    dRet = conn.returnInsertResult(iUpdate)
                    if dRet["result"] == "success":
                        return logAndRet("success", "De-Attached Successfully")
                    else:
                        return logAndRet("failure", "Couldn't De-Attach")
                else:
                    return logAndRet("failure", "Invalid Payload")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

# Remove discovered machine
def removeDiscoveredMachine(discovery_id):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = "delete from ai_device_discovery where pk_discovery_id in({0})".format(discovery_id.replace('_', ','))
                iRet = conn.returnInsertResult(sQuery)
                if iRet['result'] == 'success':
                    return logAndRet("success", "Removed discovered resources")
                else:
                    return logAndRet("failure", "Failed to remove discovered resources")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

# Remove Actual Machine
def removeActualMachine(machine_ids):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                #delete
                sQuery = "delete from ai_machine where machine_id in({0})".format(machine_ids.replace('_', ','))
                iRet = conn.returnInsertResult(sQuery)
                # delete from mongo config
                # delete from machine group
                if iRet['result'] == 'success':
                    return logAndRet("success", "Removed machines successfully")
                else:
                    return logAndRet("failure", "Failed to remove machines")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def removeActualMachineAndReinitiateDiscovery(machine_ids):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                #delete
                sQuery = "delete from ai_machine where machine_id in({0}) RETURNING ip_address".format(machine_ids.replace('_', ','))
                iRet = conn.returnSelectQueryResultWithCommit(sQuery)
                if iRet['result'] == 'success':
                    # delete from mongo config
                    # delete from machine group
                    # Start Discovery
                    ip_address = iRet['data'][0]['ip_address']
                    out = initiateRESTDiscovery({'ip_range': ip_address})
                    # out = startDiscovery({'ip_range': ip_address})
                    return logAndRet("success", "Removed and Re-discovered machines successfully")
                else:
                    return logAndRet("failure", "Failed to remove machines")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getIPAddress4Hostname(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = "select ip_address from ai_machine where lower(machine_fqdn)=lower('{0}') and active_yn='Y'".format(
                    dPayload['hostname']
                )
                dRet = conn.returnSelectQueryResult(sQuery)
                if dRet['result'] == 'success':
                    return json.dumps({'result': 'success', 'data': dRet['data'][0]['ip_address']})
                else:
                    return json.dumps({'result': 'failure', 'data': 'Unable to fetch ip address'})
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

# void
def getremotingDetialsNew(dPayload):
    try:
        lH = ['ip_address']
        lM = ['ip_address']
        if val.isPayloadValid(dPayload, lH, lM):
            sQuery = """select 
            	c.cred_type protocol, c.username, c.password, m.console_port 
            from 
            	ai_machine m inner join ai_device_credentials c on(m.fk_cred_id=c.cred_id) 
            where 
            	m.ip_address='{0}' and  c.active_yn='Y' limit 1""".format(dPayload['ip_address'])
            dRet = conn.returnSelectQueryResult(sQuery)
            if dRet['result'] == 'success':
                out = dRet['data'][0]
                password = aes.decrypt(out['password'].encode(), '@ut0!ntell!'.encode()).decode('utf-8')
                if out['protocol'].lower() == 'winrm':
                    out['protocol'] = 'rdp'

                if dRet['result'] == 'success':
                    return json.dumps(
                        {'protocol': out['protocol'].lower(), 'username': out['username'], 'password': password,
                         'port': str(out['console_port'])})
                else:
                    return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': ''})
            else:
                return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': ''})
        else:
            return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': ''})
    except Exception as e:
        return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': ''})

def getRemotingKey(dPayload):
    try:
        lH = ['ip_address', 'logged_in_user', 'alert_id']
        lM = ['ip_address', 'logged_in_user', 'alert_id']
        if val.isPayloadValid(dPayload, lH, lM):
            # Generate a random key
            import random
            import string
            an_set = string.ascii_letters + ''.join([str(i) for i in range(0, 10)])
            key = ''.join([random.choice(an_set) for i in range(0, 64)])
            videofilename = dt.now().strftime('%Y%m%d%H%M%S%f')
            l = list(videofilename)
            keyComplex = ''.join([j + l.pop(0) for j in [key[i:i + 2] for i in range(0, 64, 2)] if len(l) > 0])

            # Push Video details
            iQuery = "insert into audit_videos(alert_id, ip_address, user_id, video_name, start_datetime, created_by, created_on, key, key_active) values('{0}', '{1}', '{2}', '{3}', now(), 'system', now(), '{4}', '{5}')".format(
                dPayload['alert_id'], dPayload['ip_address'], dPayload['logged_in_user'], videofilename, keyComplex, 'Y'
            )
            dRet = conn.returnInsertResult(iQuery)
            if dRet['result'] == 'success':
                return json.dumps({'result': 'success', 'data': {'key': keyComplex}})
            return json.dumps({'result': 'failure', 'data': 'Unable to generate key'})
        else:
            return json.dumps({'result': 'failure', 'data': 'Invalid Payload'})
    except Exception as e:
        return json.dumps({'result': 'failure', 'data': 'Unable to generate key'})

# used by java client
def getRemotingDetails(dPayload):
    try:
        lH = ['key']
        lM = ['key']
        if val.isPayloadValid(dPayload, lH, lM):
            # Check for expiry
            filenameQuery = "select  ip_address, video_name from audit_videos where key='{0}' and key_active='Y'".format(dPayload['key'])
            fRet = conn.returnSelectQueryResult(filenameQuery)
            if fRet['result'] == 'failure':
                return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': '', 'video_name': '', 'ip_address': ''})

            sQuery = """select 
                        	c.cred_type protocol, c.username, c.password, m.console_port, m.platform, m.machine_fqdn 
                        from 
                        	ai_machine m inner join ai_device_credentials c on(m.fk_cred_id=c.cred_id) 
                        where 
                        	m.ip_address=(select ip_address from audit_videos where key='{0}' limit 1) and  c.active_yn='Y' limit 1""".format(
                dPayload['key']
            )
            dRet = conn.returnSelectQueryResult(sQuery)
            if dRet['result'] == 'success':
                out = dRet['data'][0]
                password = ""
                if out['protocol'].lower() != 'pam':
                    password = aes.decrypt(out['password'].encode(), '@ut0!ntell!'.encode()).decode('utf-8')
                if out['protocol'].lower() == 'winrm':
                    out['protocol'] = 'rdp'

                retPAM = {}
                if out['protocol'].lower() == 'pam':
                    if pam.lower() == 'iress':
                        retPAM = pam1.queryPassword(out['machine_fqdn'], out['username'], out['platform'], pam)
                    else:
                        retPAM = pam1.queryPassword(fRet['data'][0]['ip_address'], out['username'], out['platform'], pam)

                    print("arcon:{0}".format(retPAM))

                # Disable Key
                sFinalJSON = {}
                uQuery = "update audit_videos set key_active='Y' where key='{0}'".format(dPayload['key'])
                uRet = conn.returnInsertResult(uQuery)
                if uRet['result'] == 'success':
                    if out['protocol'].lower() == 'pam':
                        if retPAM['result'] == 'success':
                            if out['platform'].lower().__contains__('window'):
                                if retPAM['data']['username'].__contains__('\\'):
                                    sFinalJSON = json.dumps(
                                        {'protocol': 'rdp', 'domain': retPAM['data']['username'].split('\\')[0],
                                         'username': retPAM['data']['username'].split('\\')[1],
                                         'password': retPAM['data']['password'],
                                         'port': str(out['console_port']), 'video_name': fRet['data'][0]['video_name'],
                                         'ip_address': fRet['data'][0]['ip_address']})
                                else:
                                    sFinalJSON = json.dumps(
                                        {'protocol': 'rdp', 'domain': "",
                                         'username': retPAM['data']['username'],
                                         'password': retPAM['data']['password'],
                                         'port': str(out['console_port']), 'video_name': fRet['data'][0]['video_name'],
                                         'ip_address': fRet['data'][0]['ip_address']})

                            else:
                                if retPAM['data']['username'].__contains__('\\'):
                                    sFinalJSON = json.dumps(
                                        {'protocol': 'ssh', 'domain': retPAM['data']['username'].split('\\')[0],
                                         'username': retPAM['data']['username'].split('\\')[1],
                                         'password': retPAM['data']['password'],
                                         'port': str(out['console_port']), 'video_name': fRet['data'][0]['video_name'],
                                         'ip_address': fRet['data'][0]['ip_address']})
                                else:
                                    sFinalJSON = json.dumps(
                                        {'protocol': 'ssh', 'domain': "", 'username': retPAM['data']['username'],
                                         'password': retPAM['data']['password'],
                                         'port': str(out['console_port']), 'video_name': fRet['data'][0]['video_name'],
                                         'ip_address': fRet['data'][0]['ip_address']})
                            return sFinalJSON

                    if out['username'].__contains__('\\'):
                        sFinalJSON = json.dumps(
                            {'protocol': out['protocol'].lower(), 'domain': out['username'].split('\\')[0],
                             'username': out['username'].split('\\')[1], 'password': password,
                             'port': str(out['console_port']), 'video_name': fRet['data'][0]['video_name'],
                             'ip_address': fRet['data'][0]['ip_address']})
                    else:
                        sFinalJSON = json.dumps(
                            {'protocol': out['protocol'].lower(), 'domain': "", 'username': out['username'], 'password': password,
                             'port': str(out['console_port']), 'video_name': fRet['data'][0]['video_name'],
                             'ip_address': fRet['data'][0]['ip_address']})
                    print(sFinalJSON)
                    return sFinalJSON
                else:
                    return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': '', 'video_name': '', 'ip_address': '', 'domain': ''})
            else:
                return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': '', 'video_name': '', 'ip_address': '', 'domain': ''})
        else:
            return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': '', 'video_name': '', 'ip_address': '', 'domain': ''})
    except Exception as e:
        return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': '', 'video_name': '', 'ip_address': '', 'domain': ''})

def getVideos4AlertPage(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                lH = ['alert_id']
                lM = ['alert_id']
                if val.isPayloadValid(dPayload, lH, lM):
                    sQuery = "select user_id, '{0}' || video_name || '.mp4' videoname, to_char(start_datetime, 'YYYY-MM-DD HH24:MI:SS') accessed  from audit_videos where alert_id='{1}' order by start_datetime desc".format(
                        location, dPayload['alert_id']
                    )
                    dRet = conn.returnSelectQueryResultAs2DList(sQuery)
                    if dRet['result'] == 'success':
                        return json.dumps(dRet)
                    else:
                        return json.dumps({'result': 'failure', 'data': 'No sessions available'})
                else:
                    return json.dumps({'result': 'failure', 'data': 'Invalid Payload'})
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getVideos4HDDMPage(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                lH = ['ip_address']
                lM = ['ip_address']
                if val.isPayloadValid(dPayload, lH, lM):
                    sQuery = "select user_id, '{0}' || video_name || '.mp4' videoname, to_char(start_datetime, 'YYYY-MM-DD HH24:MI:SS') accessed  from audit_videos where ip_address='{1}' order by start_datetime desc".format(
                        location, dPayload['ip_address']
                    )
                    dRet = conn.returnSelectQueryResultAs2DList(sQuery)
                    if dRet['result'] == 'success':
                        return json.dumps(dRet)
                    else:
                        return json.dumps({'result': 'failure', 'data': 'no data'})
                else:
                    return json.dumps({'result': 'failure', 'data': 'Invalid Payload'})
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

# void
def getremotingDetials1(dPayload):
    try:
        lH = ['ip_address', 'logged_in_user', 'alert_id', 'videofilename']
        lM = ['ip_address', 'logged_in_user', 'alert_id', 'videofilename']
        if val.isPayloadValid(dPayload, lH, lM):

            sQuery = """select 
            	c.cred_type protocol, c.username, c.password, m.console_port 
            from 
            	ai_machine m inner join ai_device_credentials c on(m.fk_cred_id=c.cred_id) 
            where 
            	m.ip_address='{0}' and  c.active_yn='Y' limit 1""".format(dPayload['ip_address'])
            dRet = conn.returnSelectQueryResult(sQuery)
            if dRet['result'] == 'success':
                out = dRet['data'][0]
                password = aes.decrypt(out['password'].encode(), '@ut0!ntell!'.encode()).decode('utf-8')
                if out['protocol'].lower() == 'winrm':
                    out['protocol'] = 'rdp'

                # Push Video details
                iQuery = "insert into audit_videos(alert_id, ip_address, user_id, video_name, start_datetime, created_by, created_on) values('{0}', '{1}', '{2}', '{3}', now(), 'system', now())".format(
                    dPayload['alert_id'], dPayload['ip_address'], dPayload['logged_in_user'], dPayload['videofilename']
                )
                dRet = conn.returnInsertResult(iQuery)
                if dRet['result'] == 'success':
                    return json.dumps(
                        {'protocol': out['protocol'].lower(), 'username': out['username'], 'password': password,
                         'port': str(out['console_port'])})
                return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': ''})
            else:
                return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': ''})
        else:
            return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': ''})
    except Exception as e:
        return json.dumps({'protocol': '', 'username': '', 'password': '', 'port': ''})

def downloadMachineOnXLS():
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                #sQuery = "select am.machine_fqdn, am.ip_address, am.platform, am.osname, am.osversion, am.remediate, am.machine_id from ai_machine am left join ai_device_credentials ac on(am.fk_cred_id=ac.cred_id)"
                sQuery = "select inventory from ai_machine where active_yn='Y'"
                #dRet = conn.returnSelectQueryResultAs2DList(sQuery)
                dRet = conn.returnSelectQueryResult(sQuery)
                if dRet["result"] == "success":
                    xlsTmp, xlsList = dRet["data"], []
                    xlsList.append(["HOSTNAME", "IPADDRESS", "PLATFORM", "OSNAME", "VERSION", "ARCHITECTURE", "MEMORY_TOTAL", "SWAP_TOTAL", "PROCESSOR_COUNT", "DISK"])
                    for eachInv in xlsTmp:
                        inv1 = eachInv["inventory"]
                        xlRow = [inv1[i] for i in ["HOSTNAME", "IPADDRESS", "PLATFORM", "OSNAME", "VERSION", "ARCHITECTURE", "MEMORY_TOTAL", "SWAP_TOTAL", "PROCESSOR_COUNT"]] + ['\n'.join(["{0}: {1}".format(i, inv1["DISK"][i]) for i in list(inv1["DISK"].keys())])]
                        xlsList.append(xlRow)

                    row, sSys = 1, "linux"
                    xlsList[0] = [i.capitalize() for i in xlsList[0]]
                    wb = Workbook()
                    ws = wb.create_sheet("DeviceDetails")
                    thin_border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                    #Merged Header
                    ws.merge_cells('A1:J1')
                    ws.cell(row=1, column=1).value = "Autointelli Infrastructure Inventory"
                    ws['A1'].fill = PatternFill(start_color="0814FF", end_color="FFC7CE", fill_type="solid")
                    ws['A1'].font = Font(color="FFFFFF")
                    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

                    for i in xlsList:
                        ws.append(i)
                        col = 1
                        for j in i:
                            ws.cell(row=row, column=col).border = thin_border
                            col += 1
                        row += 1

                    # Header
                    for eC in "ABCDEFGHIJ":
                        ws[eC + '2'].fill = PatternFill(start_color="FFC414", end_color="FFC7CE", fill_type="solid")
                        ws[eC + '2'].alignment = Alignment(horizontal="center", vertical="center")

                    xlsxName = "AIInventory_" + str(int(datetime.now().timestamp() * 1000000)) + ".xlsx"
                    if sSys == "win":
                        xlsxPath = "E:\\" + xlsxName
                    else:
                        xlsxPath = "/usr/share/nginx/html/downloads/" + xlsxName

                    wb.remove(wb['Sheet'])
                    wb.save(xlsxPath)
                    wb.close()
                    return logAndRet("success", "http://95.216.28.228:4006/downloads/" + xlsxName)
                else:
                    return logAndRet("failure", "Unable to download machine data")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

# Return complete master data = Machine Class, Software Class, Software Sub-Class, Resource Class, Application Class, Application Sub-Class
def getMastersBasedOnName(item):
    dTableNames = {'machine_class': {'column': ['mclass_name'], 'table': 'ai_machine_class'},
                   'software_class': {'column': ['sclass_name'], 'table': 'ai_software_class'},
                   'software_subclass': {'column': ['ssclass_name'], 'table': 'ai_software_subclass'},
                   'resource_class': {'column': ['rclass_name'], 'table': 'ai_resource_class'},
                   'application_class': {'column': ['aclass_name'], 'table': 'ai_application_class'},
                   'application_subclass': {'column': ['asclass_name'], 'table': 'ai_application_subclass'}}
    sQuery = "select " + ",".join(dTableNames[item]['column']) + " from " + dTableNames[item]['table'] + " where active_yn='Y'"
    dRet = conn.returnSelectQueryResultAsList(sQuery)
    if dRet["result"] == "success":
        return json.dumps(dRet)
    else:
        return logAndRet("failure", "couldn't load the data from the database")

def getAttributes(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                type = dPayload["type"].lower()
                clas = dPayload["class"].lower()
                sub_clas = dPayload["sub_class"].lower()
                sQuery = "select attribute from attributes where lower(mars_type)='%s' and lower(item_class)='%s' and lower(item_sub_class)='%s'" %(type, clas, sub_clas)
                dRet = conn.returnSelectQueryResultAsList(sQuery)
                if dRet["result"] == "success":
                    dRet["data"]["attribute"] = dRet["data"]["attribute"][0].split(',')
                    return json.dumps(dRet)
                else:
                    return logAndRet("failure", "no data")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def createTypeEntry(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                type, clas, sub_clas,  attr = "", "", "", {}
                type = dPayload["type"]
                # The below snippet for adding software
                if type == "software":
                    name = dPayload["name"]
                    clas = dPayload["class"]
                    sub_clas = dPayload["sub_class"]
                    remediate = dPayload["remediate"]
                    attr = json.dumps(dPayload["attribute"])
                    sValQuery = "select * from ai_software where lower(software_name)='%s'" % name.lower()
                    dRet = conn.returnSelectQueryResult(sValQuery)
                    if dRet["result"] == "failure":
                        sQuery = "insert into ai_software(software_name, software_class, software_subclass, remediate, attribute, active_yn) values('%s', '%s', '%s', '%s', '%s', '%s')" % (name, clas, sub_clas, remediate, attr, 'Y')
                        dRet = conn.returnInsertResult(sQuery)
                        if dRet["result"] == "success":
                            return logAndRet("success", "Item created successfully")
                        else:
                            return logAndRet("failure", "Failed to create an entry")
                    else:
                        return logAndRet("failure", "Software Name available. Choose a different name")
                # The below snippet for adding resource
                elif type == "resource":
                    name = dPayload["name"]
                    clas = dPayload["class"]
                    sValidQuery = "select resource_id, resource_name, resource_class from ai_resource where active_yn='Y' and lower(resource_name)='%s'" % name.lower()
                    dRet = conn.returnSelectQueryResult(sValidQuery)
                    if dRet["result"] == "failure":
                        sQuery = "insert into ai_resource(resource_name, resource_class, active_yn) values('%s','%s','%s')" %(name, clas, 'Y')
                        dRet = conn.returnInsertResult(sQuery)
                        if dRet["result"] == "success":
                            return logAndRet("success", "Item created successfully")
                        else:
                            return logAndRet("failure", "Failed to create an entry")
                    else:
                        return logAndRet("failure", "Resource Name available. Choose a different name")
                # The below snippet for adding application
                elif type == "application":
                    name = dPayload["name"]
                    clas = dPayload["class"]
                    sub_clas = dPayload["sub_class"]
                    sValidQuery = "select application_id, application_name, application_class, application_subclass from ai_application where active_yn='Y' and lower(application_name)='%s'" % name.lower()
                    dRet = conn.returnSelectQueryResult(sValidQuery)
                    if dRet["result"] == "failure":
                        sQuery = "insert into ai_application(application_name, application_class, application_subclass, active_yn) values('%s','%s','%s','%s')" % (name, clas, sub_clas, 'Y')
                        dRet = conn.returnInsertResult(sQuery)
                        if dRet["result"] == "success":
                            return logAndRet("success", "Item created successfully")
                        else:
                            return logAndRet("failure", "Failed to create an entry")
                    else:
                        return logAndRet("failure", "Application Name available. Choose a different name")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def updateType(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                type, clas, sub_clas, attr = "", "", "", {}
                type = dPayload["type"]
                # The below snippet for adding software
                if type == "software":
                    name = dPayload["name"]
                    clas = dPayload["class"]
                    sub_clas = dPayload["sub_class"]
                    remediate = dPayload["remediate"]
                    attr = json.dumps(dPayload["attribute"])
                    sValQuery = "select * from ai_software where software_name='%s'" % name.lower()
                    dRet = conn.returnSelectQueryResult(sValQuery)
                    if dRet["result"] == "success":
                        sQuery = "update ai_software set software_class='%s', software_subclass='%s', remediate='%s', attribute='%s' where lower(software_name)='%s' and active_yn='Y'" %(clas, sub_clas, remediate, attr, name)
                        dRet = conn.returnInsertResult(sQuery)
                        if dRet["result"] == "success":
                            return logAndRet("success", "Item updated successfully")
                        else:
                            return logAndRet("failure", "Failed to update item")
                    else:
                        return logAndRet("failure", "Software not available to update")
                # The below snippet for adding resource
                elif type == "resource":
                    name = dPayload["name"]
                    clas = dPayload["class"]
                    sValidQuery = "select resource_id, resource_name, resource_class from ai_resource where active_yn='Y' and lower(resource_name)='%s'" % name.lower()
                    dRet = conn.returnSelectQueryResult(sValidQuery)
                    if dRet["result"] == "success":
                        sQuery = "update ai_resource set resource_class='%s' where lower(resource_name)='%s' and active_yn='%s'" % ( clas, name, 'Y')
                        dRet = conn.returnInsertResult(sQuery)
                        if dRet["result"] == "success":
                            return logAndRet("success", "Item updated successfully")
                        else:
                            return logAndRet("failure", "Failed to update item")
                    else:
                        return logAndRet("failure", "Resource not available to update")
                # The below snippet for adding application
                elif type == "application":
                    name = dPayload["name"]
                    clas = dPayload["class"]
                    sub_clas = dPayload["sub_class"]
                    sValidQuery = "select application_id, application_name, application_class, application_subclass from ai_application where active_yn='Y' and lower(application_name)='%s'" % name.lower()
                    dRet = conn.returnSelectQueryResult(sValidQuery)
                    if dRet["result"] == "success":
                        sQuery = "update ai_application set application_class='%s', application_subclass='%s' where lower(application_name)='%s' and active_yn='%s'" % (clas, sub_clas, name, 'Y')
                        dRet = conn.returnInsertResult(sQuery)
                        if dRet["result"] == "success":
                            return logAndRet("success", "Item updated successfully")
                        else:
                            return logAndRet("failure", "Failed to update item")
                    else:
                        return logAndRet("failure", "Application not available to update")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def removeType(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                type, name, sQuery = "", "", ""
                type = dPayload["type"]
                name = dPayload["name"].lower()
                if type == "software":
                    sQuery = "update ai_software set active_yn='N' where lower(software_name)='%s'" % name
                elif type == "resource":
                    sQuery = "update ai_resource set active_yn='N' where lower(resource_name)='%s'" % name
                elif type == "application":
                    sQuery = "update ai_application set active_yn='N' where lower(application_name)='%s'" % name
                dRet = conn.returnInsertResult(sQuery)
                if dRet["result"] == "success":
                    if dRet["data"] > 0:
                        return logAndRet("success", "{0} removed successfully".format(type))
                    else:
                        return logAndRet("failure", "{0} not available to delete".format(type))
                else:
                    return logAndRet("failure", "Couldn't remove {0}".format(type))
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getMARSDetails(mars_type):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = ""
                if mars_type == "software":
                    sQuery = "select software_id, software_name, software_class, software_subclass, remediate, attribute from ai_software where active_yn='Y'"
                elif mars_type == "resource":
                    sQuery = "select resource_id, resource_name, resource_class from ai_resource where active_yn='Y'"
                elif mars_type == "application":
                    sQuery = "select application_id, application_name, application_class, application_subclass from ai_application where active_yn='Y'"
                dRet = conn.returnSelectQueryResult(sQuery)
                return json.dumps(dRet)
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getMARSDetailsParticular(mars_type, mars_name):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = ""
                if mars_type == "software":
                    sQuery = "select software_id, software_name, software_class, software_subclass, remediate, attribute from ai_software where active_yn='Y' and lower(software_name)='%s'" % mars_name.lower()
                elif mars_type == "resource":
                    sQuery = "select resource_id, resource_name, resource_class from ai_resource where active_yn='Y' and lower(resource_name)='%s'" % mars_name.lower()
                elif mars_type == "application":
                    sQuery = "select application_id, application_name, application_class, application_subclass from ai_application where active_yn='Y' and lower(application_name)='%s'" % mars_name.lower()
                dRet = conn.returnSelectQueryResult(sQuery)
                return json.dumps(dRet)
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing
