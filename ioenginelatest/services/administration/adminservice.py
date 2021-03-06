#============================================================================================
# Author        : Dinesh Balaraman
# Created On    :
# (c) Copyright by Autointelli Technologies
#============================================================================================

from flask import request
import json
from services.utils import sessionkeygen
from services.utils.ConnPostgreSQL import returnInsertResult, returnSelectQueryResult, returnSelectQueryResultAsList, returnSelectQueryResultWithCommit, returnSelectQueryResultAs2DList
import services.utils.LDAPAuth as ldapauth
from services.licensing import licensing
import services.utils.ED_AES256 as aes
import services.utils.utils as ut
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from datetime import datetime
from services.utils.ConnLog import create_log_file
from services.utils.ValidatorSession import chkValidRequest, chkKeyExistsInHeader, lam_api_key_invalid, lam_api_key_missing
import services.utils.LFColors as lfc
import services.utils.mailservice as ms
import requests

#https://github.com/corydolphin/flask-cors/blob/master/examples/app_based_example.py
# Empty lambda : lambda _ : return null

lfcObj = lfc.bcolors()
CERROR, CWARN = lfcObj.printerr, lfcObj.printwar

logObj = create_log_file()
if not logObj:
    CERROR("Not able to create logfile")
    exit(0)
logERROR, logWARN, logINFO = logObj.error, logObj.warn, logObj.info

lam_api_key_missing = lam_api_key_missing()
lam_api_key_invalid = lam_api_key_invalid()

def logAndRet(s, x):
    if s == "failure":
        logERROR(x)
    else:
        logINFO(x)
    return json.dumps({"result": s, "data": x})

def getAllRoles():
    """Method: Returns the roles available in backend"""
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                rCall = returnSelectQueryResultAsList("select role_name from tbl_role where active_yn='Y'")
                return json.dumps(rCall)
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getAllTabs():
    """Method: Returns the tabs available in backend"""
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                rCall = returnSelectQueryResultAsList("select tab_name from tbl_tab where active_yn='Y'")
                return json.dumps(rCall)
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getAllPermission():
    """Method: Returns the permission available in backend"""
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                rCall = returnSelectQueryResultAsList("select permission_name from tbl_permission where active_yn='Y'")
                return json.dumps(rCall)
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getRolesMappers():
    """Method: Returns the role/tab/permission mappings available in backend"""
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = """select 
                                                    t.role_name, tab.tab_name, p.permission_name from tbl_role_tab_permission rtp, tbl_permission p, tbl_role t, tbl_tab tab 
                                            where
                                                    rtp.fk_role_id = t.pk_role_id and
                                                    rtp.fk_tab_id = tab.pk_tab_id and
                                                    rtp.fk_permission_id = p.pk_permission_id and
                                                    t.active_yn = 'Y'"""
                rCall = returnSelectQueryResult(sQuery)
                lUnpack, lPack = rCall["data"], []
                setRoleName = set([i["role_name"] for i in lUnpack])
                for i in setRoleName:
                    ltmp = [x for x in lUnpack if x["role_name"] == i]
                    oob = [y.pop("role_name") for y in ltmp]
                    lPack.append({i: [z for z in ltmp if not z.__contains__("role_name")]})
                    lUnpackTmp = [xx for xx in lUnpack if xx.__contains__("role_name")]
                    lUnpack = lUnpackTmp
                rCall["data"] = lPack
                return json.dumps(rCall)
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getOneRoleMappers(role_name):
    """Method: Returns the role/tab/permission mappings available in backend"""
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = "select role_name from tbl_role where active_yn='Y' and lower(role_name) = '{0}'".format(
                    role_name.lower().strip())
                dRet = returnSelectQueryResult(sQuery)
                if dRet["result"] == "failure":
                    return json.dumps({"result": "failure", "data": "Role '" + role_name + "' not found"})
                sQuery = """select 
                                                    t.role_name, tab.tab_name, p.permission_name from tbl_role_tab_permission rtp, tbl_permission p, tbl_role t, tbl_tab tab 
                                            where
                                                    rtp.fk_role_id = t.pk_role_id and
                                                    rtp.fk_tab_id = tab.pk_tab_id and
                                                    rtp.fk_permission_id = p.pk_permission_id and
                                                    t.role_name = '""" + role_name + "' and rtp.active_yn='Y' and t.active_yn='Y'"
                rCall = returnSelectQueryResult(sQuery)
                if rCall["result"] == "success":
                    lUnpack, lPack = rCall["data"], []
                    setRoleName = set([i["role_name"] for i in lUnpack])
                    for i in setRoleName:
                        ltmp = [x for x in lUnpack if x["role_name"] == i]
                        oob = [y.pop("role_name") for y in ltmp]
                        lPack.append({i: [z for z in ltmp if not z.__contains__("role_name")]})
                        lUnpackTmp = [xx for xx in lUnpack if xx.__contains__("role_name")]
                        lUnpack = lUnpackTmp
                    k = role_name
                    v = lPack[0][role_name]  # [0]
                    rCall["data"] = {k: v}
                    return json.dumps(rCall)
                else:
                    return logAndRet("success", "no data")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def createRole(paraJSON):
    """"Methods: Create new role based on the tabs and permission for the correspinding tabs """
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                dPayload = paraJSON  # request.get_json()
                if "role_name" in dPayload.keys() and "mapping" in dPayload.keys():
                    # Check for role_name in database
                    ret = returnSelectQueryResult("select role_name from tbl_role where role_name='" + dPayload[
                        "role_name"] + "' and active_yn='Y'")
                    if ret["result"] == "failure" and ret["data"] == "no data":
                        # Proceed inserting data, new role
                        ret = returnInsertResult(
                            "insert into tbl_role(role_name, active_yn) values('" + dPayload["role_name"] + "','Y') ")
                        if ret['result'] == 'success' and ret["data"] > 0:
                            dMapper = dPayload["mapping"]
                            iMultiBool = 0
                            for eachTab in dMapper.keys():
                                xRoleName = "select pk_role_id from tbl_role where role_name='" + dPayload[
                                    "role_name"] + "'"
                                xTabName = "select pk_tab_id from tbl_tab where tab_name='" + eachTab + "'"
                                xPermission = "select pk_permission_id from tbl_permission where permission_name='" + \
                                              dMapper[eachTab] + "'"
                                sQuery = "insert into tbl_role_tab_permission(fk_role_id, fk_tab_id, fk_permission_id, active_yn) values((" + xRoleName + "), (" + xTabName + "), (" + xPermission + ") , 'Y')"
                                ret = returnInsertResult(sQuery)
                                if ret['result'] == 'success' and ret["data"] > 0:
                                    continue
                                else:
                                    iMultiBool = 1
                                    break
                            if iMultiBool == 0:
                                return logAndRet("success", "Role {0} added successfully".format(dPayload["role_name"]))
                            else:
                                return logAndRet("failure", "Unable to add role {0}".format(dPayload["role_name"]))
                        else:
                            return logAndRet("failure", "Unable to add role {0}".format(dPayload["role_name"]))
                    else:
                        return logAndRet("failure", "role {0} already exists".format(dPayload["role_name"]))
                else:
                    return logAndRet("failure", """Payload Error - Absent of required information. 
                                    Sample Payload : {"role_name" : "x", "mapping" : {"user_management" : "RWX", "dashboard" : "RW"} }""")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def removeRole(role_name):
    """Method is used to make the role inactive and also to make all the mapping inactive. We don't delete any data instead make them inactive"""
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sChkDependecies = "select * from tbl_user_details where fk_role_id = (select pk_role_id from tbl_role where role_name='" + role_name + "') and active_yn='Y'"
                ret = returnSelectQueryResult(sChkDependecies)
                if not ret["result"] == "success":
                    sQuery = "select * from tbl_role where role_name='" + role_name + "' and active_yn='Y'"
                    ret = returnSelectQueryResult(sQuery)
                    if ret["result"] == "success":
                        if len(ret["data"]) > 0:
                            sUpdateQuery = "update tbl_role_tab_permission set active_yn='N' where fk_role_id = (select pk_role_id from tbl_role where role_name='" + role_name + "')"
                            sUpdateQuery1 = "update tbl_role set active_yn='N' where role_name='" + role_name + "'"
                            ret = returnInsertResult(sUpdateQuery)
                            ret1 = returnInsertResult(sUpdateQuery1)
                            if ret["result"] == "success" and ret1["result"] == "success":
                                if ret["data"] > 0 and ret1["data"] > 0:
                                    return logAndRet("success", "Role {0} deleted successfully".format(role_name))
                            else:
                                return logAndRet("failure", "Unable to delete the role {0}".format(role_name))
                    else:
                        return logAndRet("failure", "Role {0} not available to delete".format(role_name))
                else:
                    return logAndRet("failure", "Deleting role: {0} failed. There are users with this role.".format(role_name))
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def updateRole(role_name):
    """Method is used to make the role inactive and also to make all the mapping inactive. We don't delete any data instead make them inactive"""
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = "select * from tbl_role where role_name='" + role_name + "' and active_yn='Y'"
                ret = returnSelectQueryResult(sQuery)
                if ret["result"] == "success":
                    if len(ret["data"]) > 0:
                        dPayload = request.get_json()
                        iDQuery = "delete from tbl_role_tab_permission where fk_role_id=(select pk_role_id from tbl_role where lower(role_name)='" + role_name.lower() + "' and active_yn='Y')"
                        iDRet = returnInsertResult(iDQuery)
                        if iDRet["result"] == "success":
                            lToDB = dPayload["mapping"].keys()
                            for i in lToDB:
                                xRoleName = "select pk_role_id from tbl_role where role_name='" + role_name + "'"
                                xTabName = "select pk_tab_id from tbl_tab where tab_name='" + i + "'"
                                xPermission = "select pk_permission_id from tbl_permission where permission_name='" + \
                                              dPayload["mapping"][i] + "'"
                                iUpdate = "insert into tbl_role_tab_permission(fk_role_id, fk_tab_id, fk_permission_id, active_yn) values((" + xRoleName + "), (" + xTabName + "), (" + xPermission + ") , 'Y')"
                                iRet = returnInsertResult(iUpdate)
                            return logAndRet("success", "Role {0} updated successfully".format(role_name))
                        else:
                            return logAndRet("success", "Role {0} updated failed".format(role_name))
                else:
                    return logAndRet("failure", ret["data"])
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def updateRole1(role_name):
    """Method is used to make the role inactive and also to make all the mapping inactive. We don't delete any data instead make them inactive"""
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = "select * from tbl_role where role_name='" + role_name + "' and active_yn='Y'"
                ret = returnSelectQueryResult(sQuery)
                if ret["result"] == "success":
                    if len(ret["data"]) > 0:
                        dPayload = request.get_json()
                        sAvailQuery = "select distinct t.tab_name from tbl_role_tab_permission rtp, tbl_tab t where rtp.fk_tab_id = t.pk_tab_id and rtp.fk_role_id = (select pk_role_id from tbl_role where role_name='" + role_name + "')";
                        rAvailTabs = returnSelectQueryResult(sAvailQuery)
                        if rAvailTabs["result"] == "success":
                            lFromDB = [i["tab_name"] for i in rAvailTabs["data"]]
                            lToDB = dPayload["mapping"].keys()
                            lMakeInactive = [i for i in lFromDB if not i in lToDB]
                            for i in lMakeInactive:
                                iUpdate = "update tbl_role_tab_permission set active_yn = 'N' where fk_role_id = (select pk_role_id from tbl_role where role_name='" + role_name + "' and active_yn='Y') and fk_tab_id = (select pk_tab_id from tbl_tab where tab_name='" + i + "')"
                                iRet = returnInsertResult(iUpdate)
                            lMakeUpdate = [i for i in lToDB if i in lFromDB]
                            for i in lMakeUpdate:
                                sPerm = dPayload["mapping"][i]
                                sTabname = i
                                iUpdate = "update tbl_role_tab_permission set fk_permission_id=(select pk_permission_id from tbl_permission where permission_name='" + sPerm + "') where fk_role_id=(select pk_role_id from tbl_role where role_name='" + role_name + "') and fk_tab_id=(select pk_tab_id from tbl_tab where tab_name='" + sTabname + "')"
                                iRet = returnInsertResult(iUpdate)
                            lMakeInsert = [i for i in lToDB if not i in lFromDB]
                            for i in lMakeInsert:
                                xRoleName = "select pk_role_id from tbl_role where role_name='" + role_name + "'"
                                xTabName = "select pk_tab_id from tbl_tab where tab_name='" + i + "'"
                                xPermission = "select pk_permission_id from tbl_permission where permission_name='" + \
                                              dPayload["mapping"][i] + "'"
                                iUpdate = "insert into tbl_role_tab_permission(fk_role_id, fk_tab_id, fk_permission_id, active_yn) values((" + xRoleName + "), (" + xTabName + "), (" + xPermission + ") , 'Y')"
                                iRet = returnInsertResult(iUpdate)
                            return logAndRet("success", "Role {0} updated successfully".format(role_name))
                else:
                    return logAndRet("failure", ret["data"])
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def acceptSessionKey(dPayload):
    try:
        pSessKey, pUserID = dPayload['sk'], dPayload['uid']
        sQuery = "insert into tbl_session_keys(session_key, active_yn, fk_user_id) values('{0}','Y',{1})".format(pSessKey, pUserID)
        dRet = returnInsertResult(sQuery)
        if dRet["result"] == "success":
            if dRet["data"] > 0:
                return logAndRet('success', 'received key from central instance')
    except Exception as e:
        return logAndRet('success', 'received key from central instance. failed to register')

def push2AllLocation(pSession, pUID):
    try:
        p, f = 0, 0
        for i in ['r2d22', 'r2d23', 'r2d21']:
            sURL = "https://{0}.nxtgen.com/ui/api1.0/loginregi".format(i)
            dHeader = {'Content-Type': 'application/json'}
            jsonData = {'sk': pSession, 'uid': pUID}
            ret = requests.post(url=sURL, headers=dHeader, json=jsonData, verify=False)
            if ret.status_code == 200:
                p += 1
            else:
                f += 1
            logINFO('Share session: success: {0}, failed: {1}'.format(p, f))
    except Exception as e:
        logERROR('Share session: Exception: {0} '.format(str(e)))

def checkLogin(paraJSON):
    """Methods is used to check if the login is valid and return back the 128 bit key which is used as sesssion id
       payload: {"username": "admin", "password": "xxx"} """
    try:
        sLExp = ""
        #License Expiry
        bLicense = licensing.checkExpiry()
        if bLicense == False:
            sLExp = "License Expired"
            logINFO(sLExp)
            #return json.dumps({'result': 'failure', 'data': 'License Expired'})

        #Unpack Payload
        dPayload = paraJSON
        sUserName = dPayload["username"]
        sPassword = dPayload["password"]
        # dPayload["password"] if (sUserName.__contains__("\\") or sUserName.__contains__("@")) else decoder.encode('@ut0!ntell!', dPayload["password"])

        #Make authentication decision based on user type
        sUTQuery = """
select 
    ut.user_type_desc as user_type 
from 
    tbl_user_details ud, tbl_user_type ut 
where 
    ud.fk_user_type = ut.pk_user_type_id and ud.user_id='""" + sUserName.strip() + """'"""
        dRet = returnSelectQueryResult(sUTQuery)
        if dRet["result"] == "success":
            if dRet["data"][0]["user_type"].upper() == "LDAP":

                # Call LDAP Auth & Client Decryption
                k = '@ut0!ntell!'.encode()
                fromClient = sPassword.encode()
                pass_de = aes.decrypt(fromClient, k).decode('utf-8')
                # BUG for \r and \n in username of AD
                x = ""
                if sUserName.find('\r'):
                    x = sUserName.replace('\r', '\\r')
                elif sUserName.find('\n'):
                    x = sUserName.replace('\n', '\\n')
                dRetLDAP = ldapauth.authenticate(x.split('@')[0], pass_de)
                if dRetLDAP["result"] == "success":

                    # Once the LDAP auth is success, send user information
                    dFinalData = {}
                    sQuery = """
                                                        select 
                                                        	ud.pk_user_details_id, ud.user_id, ud.first_name, ud.last_name, ud.email_id, ud.phone_number, z.country_code, z.time_zone, z.gmt_offset, r.role_name, ud.first_time_login, ut.user_type_desc user_type  
                                                        from 
                                                        	tbl_user_details ud, tbl_role r, tbl_zone z, tbl_user_type ut 
                                                        where 
                                                        	ud.fk_role_id = r.pk_role_id and
                                                        	ud.fk_time_zone_id = z.pk_zone_id and 
                                                        	ud.fk_user_type = ut.pk_user_type_id and 
                                                        	ud.user_id='""" + sUserName + """' and 
                                                        	ud.active_yn='Y'"""
                    sRet = returnSelectQueryResult(sQuery)
                    if sRet["result"] == "success":
                        sRet['data'][0]['aiorch_user_id'] = sRet['data'][0]['user_id'].split('@')[0] #Orchestration id
                        dFinalData.update(sRet['data'][0])  # UserInfo

                        #Attach Session Key
                        dRetKey = sessionkeygen.createSession(str(sRet['data'][0]['pk_user_details_id']))
                        dFinalData['session_id'] = dRetKey['data']
                        # push2AllLocation(dFinalData['session_id'], 1)
                        if dRetKey["result"] == "success":
                            # Send role details along to customize UI
                            pRoleName = sRet['data'][0]['role_name']
                            dRetMap = getRoleMappers(pRoleName)
                            dFinalData['mapper'] = dRetMap['data']
                            dFinalData["result"] = "success"
                            dFinalData['license'] = sLExp
                            logINFO("user {0} logged in with LDAP auth".format(sUserName))
                            return json.dumps(dFinalData)
                        else:
                            return logAndRet("failure", "Failed accepting request. Try after sometime.")
                    else:
                        return logAndRet("failure", "Failed fetching profile. Try after sometime.")
                else:
                    return logAndRet("failure", dRetLDAP["data"])

            else:

                #Non-LDAP Users
                dFinalData = {}
                sQuery = """
                select 
                	ud.pk_user_details_id, ud.user_password, ud.user_id, ud.first_name, ud.last_name, ud.email_id, ud.phone_number, z.country_code, z.time_zone, z.gmt_offset, r.role_name, ud.first_time_login, ut.user_type_desc user_type  
                from 
                	tbl_user_details ud, tbl_role r, tbl_zone z, tbl_user_type ut 
                where 
                	ud.fk_role_id = r.pk_role_id and
                	ud.fk_time_zone_id = z.pk_zone_id and 
                	ud.fk_user_type = ut.pk_user_type_id and 
                	ud.user_id='""" + sUserName + """' and  
                	ud.active_yn='Y'"""
                #ud.user_password='""" + sPassword + """' and
                sRet = returnSelectQueryResult(sQuery)
                if sRet["result"] == "success":

                    if sUserName.lower() != "admin":
                        #Check for Account Lock
                        sQuery = "select COALESCE(attempts,0,attempts) attempts from tbl_user_details where user_id='%s' and active_yn='Y'" % sUserName
                        accRet = returnSelectQueryResult(sQuery)
                        if accRet['result'] == 'success':
                            iAttempts = int(accRet['data'][0]['attempts'])
                            if iAttempts >= 3:
                                return json.dumps({'result': 'failure', 'data': 'Account locked! Contact Administrator.'})

                    #Compare client and DB encrypted password
                    k = '@ut0!ntell!'.encode()
                    fromClient = sPassword.encode()
                    cmp1 = aes.decrypt(fromClient, k).decode('utf-8')
                    fromDB = sRet['data'][0]['user_password'].encode()
                    cmp2 = aes.decrypt(fromDB, k).decode('utf-8')
                    if cmp1 == cmp2:

                        #Send User Info
                        del sRet['data'][0]['user_password']
                        sRet['data'][0]['aiorch_user_id'] = sRet['data'][0]['user_id'].split('@')[0]  # Orchestration id
                        dFinalData.update(sRet['data'][0])  # UserInfo
                        #Attach Session key
                        dRetKey = sessionkeygen.createSession(str(sRet['data'][0]['pk_user_details_id']))
                        dFinalData['session_id'] = dRetKey['data']
                        # push2AllLocation(dFinalData['session_id'], 1)
                        if dRetKey["result"] == "success":
                            # Send role details along to customize UI
                            pRoleName = sRet['data'][0]['role_name']
                            dRetMap = getRoleMappers(pRoleName)
                            dFinalData['mapper'] = dRetMap['data']
                            dFinalData["result"] = "success"
                            dFinalData['license'] = sLExp
                            logINFO("user {0} logged in with Non-LDAP auth".format(sUserName))
                            return json.dumps(dFinalData)
                        else:
                            return logAndRet("failure", "Failed accepting request. Try after sometime.")
                    else:
                        #Increase failed attempts by 1
                        iQuery = "update tbl_user_details set attempts=attempts+1 where user_id='%s'" % sUserName
                        dRet = returnInsertResult(iQuery)
                        return logAndRet("failure", 'Wrong Password! entered by {0}'.format(sUserName))
                else:
                    return logAndRet("failure", "User {0} doesn't exists".format(sUserName))
        else:
            return logAndRet("failure", "User {0} doesn't exists".format(sUserName))

    except Exception as e:
        return logAndRet("failure", "Exception: {0}".format(str(e)))

def firstTimeLoginPwdReset():
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                dPayload = request.get_json()
                lAttr = ["pk_user_details_id", "user_password"]
                lPayErr = [1 if i in lAttr else 0 for i in dPayload.keys()]
                if not 0 in lPayErr:
                    k = '@ut0!ntell!'.encode()
                    fromClient = dPayload["user_password"].encode()
                    received_pwd = aes.decrypt(fromClient, k).decode('utf-8')
                    to_db = aes.encrypt(received_pwd.encode(), k).decode('utf-8')
                    iQuery = "update tbl_user_details set user_password='{0}', first_time_login='N' where pk_user_details_id={1}".format(to_db, dPayload["pk_user_details_id"])
                    dRet = returnInsertResult(iQuery)
                    if dRet["result"] == "success":
                        return json.dumps({"result": "success", "data": "First Time Password Set"})
                    else:
                        return json.dumps({"result": "failure", "data": "Failed to Set First Time Password"})
                else:
                    return logAndRet("failure", """Payload Error - Absent of required information. 
                                                            Sample Payload : {"pk_user_details_id": , "user_password": ""} """)
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
            return lam_api_key_missing


def getRoleMappers(role_name):
    """Method: Returns the role/tab/permission mappings available in backend"""
    try:
        sQuery = """select 
                                        t.role_name, tab.tab_name, p.permission_name from tbl_role_tab_permission rtp, tbl_permission p, tbl_role t, tbl_tab tab 
                                where
                                        rtp.fk_role_id = t.pk_role_id and
                                        rtp.fk_tab_id = tab.pk_tab_id and
                                        rtp.fk_permission_id = p.pk_permission_id and 
                                        rtp.active_yn = 'Y' and 
                                        t.role_name = '""" + role_name + "'"
        rCall = returnSelectQueryResult(sQuery)
        if rCall["result"] == "success":
            lUnpack, lPack = rCall["data"], []
            setRoleName = set([i["role_name"] for i in lUnpack])
            for i in setRoleName:
                ltmp = [x for x in lUnpack if x["role_name"] == i]
                oob = [y.pop("role_name") for y in ltmp]
                lPack.append({i: [z for z in ltmp if not z.__contains__("role_name")]})
                lUnpackTmp = [xx for xx in lUnpack if xx.__contains__("role_name")]
                lUnpack = lUnpackTmp
            k = role_name
            v = lPack[0][role_name]
            rCall["data"] = {k: v}
            return rCall
        else:
            return logAndRet("success", "Role '" + role_name + "' not found")
    except Exception as e:
        return logAndRet("failure", "Exception: {0}".format(str(e)))

def logout():
    """Method is used to logout the user and destroy the session key"""
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                key = request.headers["SESSIONKEY"]
                sQuery = "update tbl_session_keys set active_yn='N' where session_key='" + key.strip() + "'"
                dRet = returnInsertResult(sQuery)
                if dRet["result"] == "success":
                    logINFO("session {0} logged out".format(key))
                    return json.dumps({"result": "success", "data": "Successfully logged out"})
                else:
                    return logAndRet("failure", "Unable to logout")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getAllTimeZone():
    "This methods gets all the time zone details"
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = "select pk_zone_id, country_code, time_zone, gmt_offset from tbl_zone where active_yn='Y'"
                dRet = returnSelectQueryResult(sQuery)
                if dRet["result"] == "success":
                    return json.dumps(dRet)
                else:
                    return logAndRet("failure", "no data")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getUserTypes():
    "This methods is used to get the user types"
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = "select user_type_desc from tbl_user_type where active_yn='Y'"
                dRet = returnSelectQueryResultAsList(sQuery)
                if dRet["result"] == "success":
                    return json.dumps(dRet)
                else:
                    return logAndRet("failure", "no data")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getUserIDs():
    "This methods is used to get the user types"
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = "select user_id from tbl_user_details where active_yn='Y'"
                dRet = returnSelectQueryResultAsList(sQuery)
                if dRet["result"] == "success":
                    return json.dumps(dRet)
                else:
                    return logAndRet("failure", "no data")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getUserCreationOneTimeData():
    "This methods gets the details those are required to create a user. existing user ids, zone details, roles, user types."
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                dPack = {}
                dPack["roles"] = json.loads(getAllRoles())["data"]
                dPack["zones"] = json.loads(getAllTimeZone())["data"]
                dPack["userids"] = json.loads(getUserIDs())["data"]
                dPack["usertypes"] = json.loads(getUserTypes())["data"]
                dFinalOutput = {"result": "success", "data": dPack}
                return json.dumps(dFinalOutput)
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getSingleUser(user_name):
    "This method fetches all the users from the database"
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = """select 
    	user_id, first_name, middle_name, last_name, email_id, phone_number, role.role_name, zon.pk_zone_id zone_id, zon.country_code || ' - ' || zon.time_zone as TimeZone, usrty.user_type_desc  
    from 
    	tbl_user_details usr, tbl_role role, tbl_zone zon, tbl_user_type usrty
    where
    	usr.fk_role_id = role.pk_role_id and
    	usr.fk_time_zone_id = zon.pk_zone_id and
    	usr.fk_user_type = usrty.pk_user_type_id and
    	usr.active_yn='Y' and usr.user_id='""" + user_name + """'"""
                dRet = returnSelectQueryResult(sQuery)
                if dRet["result"] == "success":
                    return json.dumps(dRet)
                else:
                    return logAndRet("failure", "no data")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getUsers():
    "This method fetches all the users from the database"
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                sQuery = """select 
	usr.pk_user_details_id pk_id, user_id, first_name, middle_name, last_name, email_id, phone_number, role.role_name, zon.pk_zone_id zone_id, zon.country_code || ' - ' || zon.time_zone as TimeZone  ,usrty.user_type_desc  
from 
	tbl_user_details usr, tbl_role role, tbl_zone zon, tbl_user_type usrty
where
	usr.fk_role_id = role.pk_role_id and
	usr.fk_time_zone_id = zon.pk_zone_id and
	usr.fk_user_type = usrty.pk_user_type_id and
	usr.active_yn='Y'"""
                dRet = returnSelectQueryResult(sQuery)
                if dRet["result"] == "success":
                    return json.dumps(dRet)
                else:
                    return logAndRet("failure", "no data")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def getUsersAdminSettings(user_id):
    "This method fetches all the users from the database"
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                cond = ""
                if user_id != "admin":
                    cond = " and user_id != 'admin'"
                sQuery = """select 
	usr.pk_user_details_id pk_id, user_id, first_name, middle_name, last_name, email_id, phone_number, role.role_name, zon.pk_zone_id zone_id, zon.country_code || ' - ' || zon.time_zone as TimeZone  ,usrty.user_type_desc  
from 
	tbl_user_details usr, tbl_role role, tbl_zone zon, tbl_user_type usrty
where
	usr.fk_role_id = role.pk_role_id and
	usr.fk_time_zone_id = zon.pk_zone_id and
	usr.fk_user_type = usrty.pk_user_type_id and
	usr.active_yn='Y'""" + cond
                dRet = returnSelectQueryResult(sQuery)
                if dRet["result"] == "success":
                    return json.dumps(dRet)
                else:
                    return logAndRet("failure", "no data")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def chkUserExists(userid):
    try:
        sQuery = "select * from tbl_user_details where lower(user_id)=lower('{0}')".format(userid)
        dRet = returnSelectQueryResult(sQuery)
        if dRet['result'] == 'success':
            return True
        else:
            return False
    except Exception as e:
        return True

def createUser():
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                dPayload = request.get_json()
                lAttr = ["user_id", "first_name", "middle_name", "last_name", "email_id", "phone_number", "time_zone", "role", "user_type"]
                lPayErr = [1 if i in lAttr else 0 for i in dPayload.keys()]
                if not 0 in lPayErr:
                    lMandat = ["user_id", "first_name", "last_name", "email_id", "phone_number", "time_zone", "role", "user_type"]
                    lMandatChk = [1 if dPayload[i].strip() != "" else 0 for i in lMandat]
                    if not 0 in lMandatChk:
                        if chkUserExists(dPayload["user_id"]):
                            return json.dumps({'result': 'failure', 'data': 'User already exists. Choose a different userid'})
                        pUserID = dPayload["user_id"]
                        #pPassword = decoder.decode(sDecodeKey, dPayload["user_password"])
                        #pPassword = dPayload["user_password"] #decoder.encode('@ut0!ntell!', dPayload["user_password"])
                        defaultPwd = "NxtGen@123".encode()
                        k = '@ut0!ntell!'.encode()
                        pPassword = aes.encrypt(defaultPwd, k).decode('utf-8')
                        pFirstname = dPayload["first_name"]
                        pMiddlename = dPayload["middle_name"]
                        pLastname = dPayload["last_name"]
                        pEmailID = dPayload["email_id"]
                        pPhonenumber = dPayload["phone_number"]
                        pTimezn = int(dPayload["time_zone"])
                        rt = returnSelectQueryResultAsList("select pk_role_id from tbl_role where role_name='" + dPayload["role"] + "' and active_yn='Y'")
                        pRole = rt["data"]["pk_role_id"][0]
                        #pRole = int(dPayload["role"])
                        rt = returnSelectQueryResultAsList("select pk_user_type_id from tbl_user_type where user_type_desc='" + dPayload["user_type"] + "' and active_yn='Y'")
                        pUsertyp = rt["data"]["pk_user_type_id"][0]
                        #pUsertyp = int(dPayload["user_type"])
                        flag = 'Y'
                        first_time_login = "N" if dPayload["user_type"].strip().lower() == "ldap" else "Y"

                        #LDAPAuth Chk
                        # if dPayload["user_type"].strip().lower() == "ldap":
                        #     #pUserID = pUserID if pUserID.find('@') >= 0 else ('@' + pEmailID.split('@')[1]) #"@autointellidev.com"
                        #     #if pUserID.find('@') >= 0:
                        #     #    pUserID = pUserID[pUserID.find('@')+1:][:pUserID[pUserID.find('@')+1:].find('.')] + "\\" + pUserID.split('@')[0]
                        #     #elif not pUserID.find('\\') >= 0:
                        #     #    pUserID = pEmailID[pEmailID.find('@')+1:][:pEmailID[pEmailID.find('@')+1:].find('.')] + pUserID
                        #     dLDRet = ldapauth.authenticate(pUserID, ut.decodeAuto(dPayload["user_password"]))
                        #     if dLDRet["result"] == "success":
                        #         logINFO("LDAP Auth success, {0}".format(pUserID))
                        #         pass
                        #     else:
                        #         return logAndRet("failure", dLDRet["data"])

                        # After LDAP pass, insert the record in database
                        orch_pass = 'jGl25bVBBBW96Qi9Te4V37Fnqchz/Eu4qB9vKrRIqRg='
                        iInsertQuery = """insert into tbl_user_details(user_id, user_password, first_name, middle_name, last_name, email_id, phone_number, fk_time_zone_id, fk_role_id, fk_user_type, created_by, created_on, active_yn, orch_pass, first_time_login) 
                        values('%s','%s','%s','%s','%s','%s','%s',%d,%d,%d,%d,%s,'%s','%s', '%s') RETURNING pk_user_details_id""" %(pUserID, pPassword, pFirstname, pMiddlename, pLastname, pEmailID, pPhonenumber, pTimezn, pRole, pUsertyp, 1, 'now()', flag, orch_pass, first_time_login)
                        #iRet = returnInsertResult(iInsertQuery)
                        iRet = returnSelectQueryResultWithCommit(iInsertQuery)
                        if iRet["result"] == "success":
                            msub = "New User Registration"
                            mto = [pEmailID]
                            mcc = [pEmailID]
                            mbody = "Thanks for registering to https://r2d2.nxtgen.com/nxtgen <BR/>User ID: {0}<BR/>Use this password, <B>NxtGen@123</B> to do first time login".format(
                                pUserID)
                            mlbody = "Thanks for registering to https://r2d2.nxtgen.com/nxtgen <BR/>User ID: {0}<BR/>Password: AD Password".format(
                                pUserID)
                            try:
                                if first_time_login == 'N':
                                    ms.sendmail(msub, mto, mcc, mlbody)
                                else:
                                    ms.sendmail(msub, mto, mcc, mbody)
                            except Exception as e:
                                return logAndRet("failure", "Welcome Mail Send Failed: with Exception: {0}".format(str(e)))
                            #return json.dumps({"result": "success", "data": "User added successfully"})
                            return logAndRet("success", iRet["data"][0]["pk_user_details_id"])
                        else:
                            return logAndRet("failure", "Unable to add user")
                    else:
                        return logAndRet("failure", "Mandatory fields are empty.")
                else:
                    return logAndRet("failure", """Payload Error - Absent of required information. 
                                        Sample Payload : {"user_id": "",  "user_password": "",  "first_name": "",  "middle_name": "",  "last_name": "",  "email_id": "",  "phone_number": "",  "time_zone": "",  "role": "",  "user_type": ""}""")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def deleteUser(user_name):
    """Method is used to logout the user and destroy the session key"""
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                key = request.headers["SESSIONKEY"]
                sQuery = "update tbl_user_details set active_yn='N', user_id = user_id || '.bkp' where user_id='" + user_name + "'"
                dRet = returnInsertResult(sQuery)
                if dRet["result"] == "success":
                    return logAndRet("success", "User has been removed successfully")
                else:
                    return logAndRet("failure", "Unable to remove user")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def modifyUser(pUserName, dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                #dPayload = request.get_json()
                lAttr = ["first_name", "middle_name", "last_name", "email_id", "phone_number", "time_zone", "role", "user_type"]
                lPayErr = [1 if i in lAttr else 0 for i in dPayload.keys()]
                if not 0 in lPayErr:
                    lMandat = ["first_name", "last_name", "email_id", "phone_number", "time_zone", "role", "user_type"]
                    lMandatChk = [1 if dPayload[i].strip() != "" else 0 for i in lMandat]
                    if not 0 in lMandatChk:
                        pUserID = pUserName
                        #pPassword = decoder.decode(sDecodeKey, dPayload["user_password"])
                        #pPassword = dPayload["user_password"] #decoder.encode('@ut0!ntell!', dPayload["user_password"])
                        pFirstname = dPayload["first_name"]
                        pMiddlename = dPayload["middle_name"]
                        pLastname = dPayload["last_name"]
                        pEmailID = dPayload["email_id"]
                        pPhonenumber = dPayload["phone_number"]
                        pTimezn = int(dPayload["time_zone"])
                        rt = returnSelectQueryResultAsList("select pk_role_id from tbl_role where role_name='" + dPayload["role"] + "' and active_yn='Y'")
                        pRole = rt["data"]["pk_role_id"][0]
                        #pRole = int(dPayload["role"])
                        rt = returnSelectQueryResultAsList("select pk_user_type_id from tbl_user_type where user_type_desc='" + dPayload["user_type"] + "' and active_yn='Y'")
                        pUsertyp = rt["data"]["pk_user_type_id"][0]
                        #pUsertyp = int(dPayload["user_type"])
                        flag = 'Y'

                        #LDAPAuth Chk
                        # if dPayload["user_type"].strip().lower() == "ldap":
                        #     #pUserID += pEmailID.split('@')[1] #"@autointellidev.com"
                        #     #dLDRet = ldapauth.authenticate(pUserID, aes.decrypt('@ut0!ntell!',dPayload["user_password"]).decode('utf-8'))
                        #     dLDRet = ldapauth.authenticate(pUserID, ut.decodeAuto(dPayload["user_password"]))
                        #     if dLDRet["result"] == "success":
                        #         logINFO("LDAP Auth success, {0}".format(pUserID))
                        #         pass
                        #     else:
                        #         return logAndRet("failure", dLDRet["data"])

                        # After LDAP pass, insert the record in database
                        iUpdateQuery = """update tbl_user_details set first_name='%s', middle_name='%s', last_name='%s', email_id='%s', phone_number='%s', fk_time_zone_id=%d, fk_role_id=%d, fk_user_type=%d, modified_by=%d, modified_on='%s', active_yn='%s' where user_id='%s'""" %(
                            pFirstname, pMiddlename, pLastname, pEmailID, pPhonenumber, pTimezn, pRole, pUsertyp, 1, 'now()', flag, pUserName)
                        iRet = returnInsertResult(iUpdateQuery)
                        if iRet["result"] == "success":
                            return logAndRet("success", "User updated successfully")
                        else:
                            return logAndRet("failure", "Unable to update user")
                    else:
                        return logAndRet("failure", "Mandatory fields are empty.")
                else:
                    return logAndRet("failure", """Payload Error - Absent of required information. 
                                        Sample Payload : {"user_password": "",  "first_name": "",  "middle_name": "",  "last_name": "",  "email_id": "",  "phone_number": "",  "time_zone": "",  "role": "",  "user_type": ""}""")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

'''
# These are VOID. Developed for RIL
def autoclassify():
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            sQuery = """select interaction, to_char(open_time, 'DD/MM/YYYY, HH24:MI:SS') open_time, description, category, sub_category, area, sub_area, assignment_group, status, learning from ai_auto_classify_new order by open_time desc"""
            dRet = returnSelectQueryResult(sQuery)
            return json.dumps(dRet)
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing

def updateLearning(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            uQuery = "update historical_data set category='{0}',active_yn='{1}' where pk_id={2}".format(dPayload["category"], dPayload["active_yn"], dPayload["pk_id"])
            dRet = returnInsertResult(uQuery)
            return json.dumps(dRet)
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing
'''

"""
#This is for RBC POC
def rbc_call():
    sQuery = '''
select 
        seq, ordr, app_code, user_id, role, application_name, ticket_id, status 
from 
        transaction  
where 
        active_yn='Y' 
order by seq desc, ordr asc'''
    dRet = returnSelectQueryResultAs2DList(sQuery)
    if dRet["result"] == "success":
        payload = dRet["data"]
        d = {}
        for i in payload[1:]:
            if i[0] in list(d.keys()):
                l = [i[5], i[6], i[7]]
                d[i[0]]["data"].append(l)
            else:
                l = [i[5], i[6], i[7]]
                dd = {"app_code": i[2], "user_id": i[3], "role": i[4], "data": [l]}
                d[i[0]] = dd
        print(d)
        #out = sorted(d.items(), key=lambda x: x[1], reverse=True)
        #print(out)
        out = sorted(d.items(), key=lambda x: [0], reverse=True)
        #print(out)
        out1 = [j[1] for j in out]
        print(out1)
    return json.dumps(out1)
"""

'''
#Bulk update for RIL
def bulkUpdate(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            for i in dPayload["data"]:
                uQuery = "update historical_data set category='{0}',active_yn='{1}' where pk_id={2}".format(i["category"], i["active_yn"], i["pk_id"])
                dRet = returnInsertResult(uQuery)
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing
'''

'''
#Classification load for RIL
def loadClassificationMasterData():
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            lFinal = []
            sCategory = "select distinct category from historical_data"
            sSubCategory = "select distinct sub_category from historical_data"
            sArea = "select distinct area from historical_data"
            sSubArea = "select distinct sub_area from historical_data"
            sAssGrp = "select distinct assignment_group from historical_data"
            dRet1 = returnSelectQueryResultAsList(sCategory)
            dRet2 = returnSelectQueryResultAsList(sSubCategory)
            dRet3 = returnSelectQueryResultAsList(sArea)
            dRet4 = returnSelectQueryResultAsList(sSubArea)
            dRet5 = returnSelectQueryResultAsList(sAssGrp)
            lFinal.append(dRet1["data"])
            lFinal.append(dRet2["data"])
            lFinal.append(dRet3["data"])
            lFinal.append(dRet4["data"])
            lFinal.append(dRet5["data"])
            return json.dumps({"result": "success", "data": lFinal})
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing
'''

'''
#Learning for RIL
def insertLearning(dPayload):
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            sQuery = "insert into historical_data(category, sub_category, area, sub_area, status, assignment_group, mail_description, active_yn) values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}', '{6}', '{7}')".format(
        dPayload["category"],
        dPayload["sub_category"],
        dPayload["area"],
        dPayload["sub_area"],
        "closed",
        dPayload["assignment_group"],
        dPayload["mail_description"],
        "Y")
            dRet = returnInsertResult(sQuery)
            return json.dumps({"result": "success", "data": "data added"})
'''

def downloadMachineOnXLS():
    if chkKeyExistsInHeader("SESSIONKEY"):
        if chkValidRequest(request.headers["SESSIONKEY"]):
            try:
                #sQuery = "select am.machine_fqdn, am.ip_address, am.platform, am.osname, am.osversion, am.remediate, am.machine_id from ai_machine am left join ai_device_credentials ac on(am.fk_cred_id=ac.cred_id)"
                sQuery = "select pk_id id, ticket_id, category, sub_category, area, sub_area, status, assignment_group, mail_description, modified_date, modifed_by, active_yn from historical_data where modifed_by is not null"
                #dRet = conn.returnSelectQueryResultAs2DList(sQuery)
                dRet = returnSelectQueryResult(sQuery)
                if dRet["result"] == "success":
                    xlsTmp, xlsList = dRet["data"], []
                    xlsList.append(["id", "ticket_id", "category", "sub_category", "area", "sub_area", "status", "assignment_group", "mail_description", "modified_date", "modifed_by", "active_yn"])
                    for eachInv in xlsTmp:
                        #inv1 = eachInv["inventory"]
                        #xlRow = [inv1[i] for i in ["id", "ticket_id", "category", "sub_category", "area", "sub_area", "status", "assignment_group", "mail_description", "modified_date", "modifed_by", "active_yn"]]
                        xlRow = [eachInv for i in
                                 ["id", "ticket_id", "category", "sub_category", "area", "sub_area", "status",
                                  "assignment_group", "mail_description", "modified_date", "modifed_by", "active_yn"]]
                        xlsList.append(xlRow)

                    row, sSys = 1, "linux"
                    xlsList[0] = [i.capitalize() for i in xlsList[0]]
                    wb = Workbook()
                    ws = wb.create_sheet("LearningDetails")
                    thin_border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                    #Merged Header
                    ws.merge_cells('A1:J1')
                    ws.cell(row=1, column=1).value = "Learning Details"
                    ws['A1'].fill = PatternFill(start_color="0814FF", end_color="FFC7CE", fill_type="solid")
                    ws['A1'].font = Font(color="FFFFFF")
                    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

                    for i in xlsList:
                        ws.append(i)
                        col = 1
                        for j in i:
                            ws.cell(row=row, column=col).border = thin_border
                            col += 1
                        row += 1

                    # Header
                    for eC in "ABCDEFGHIJKL":
                        ws[eC + '2'].fill = PatternFill(start_color="FFC414", end_color="FFC7CE", fill_type="solid")
                        ws[eC + '2'].alignment = Alignment(horizontal="center", vertical="center")

                    xlsxName = "AIInventory_" + str(int(datetime.now().timestamp() * 1000000)) + ".xlsx"
                    if sSys == "win":
                        xlsxPath = "E:\\" + xlsxName
                    else:
                        xlsxPath = "/usr/share/nginx/html/downloads/" + xlsxName

                    wb.remove(wb['Sheet'])
                    wb.save(xlsxPath)
                    wb.close()
                    return logAndRet("success", "http://172.16.1.35:8000/downloads/" + xlsxName)
                else:
                    return logAndRet("failure", "Unable to download machine data")
            except Exception as e:
                return logAndRet("failure", "Exception: {0}".format(str(e)))
        else:
            return lam_api_key_invalid
    else:
        return lam_api_key_missing



